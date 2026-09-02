#!/usr/bin/env python3
"""Format EPD meeting protocol: plan block, tasks, themed discussion."""

from __future__ import annotations

import re
import zipfile
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
SRC_NAME = "Протокол встреч по проекту ЭПД  ЦП-ЦКС-ЦСВ.xlsx"
OUT_PATH = ROOT / SRC_NAME
FIXED_TMP = Path("/tmp/protocol_fixed.xlsx")

# Colors
BLUE = "1F4E79"
BLUE_MID = "2F69C7"
BLUE_LIGHT = "D6E3F0"
BLUE_PALE = "EEF4FA"
GREEN = "548235"
GREEN_LIGHT = "E2EFD9"
ORANGE = "C65911"
ORANGE_LIGHT = "FCE4D6"
YELLOW = "FFF2CC"
YELLOW_HDR = "FFE699"
GRAY = "7F7F7F"
GRAY_LIGHT = "F2F2F2"
WHITE = "FFFFFF"
RED_SOFT = "FCE4EC"
TASK_FILL = "FFF8E7"
THESIS_FILL = "F5F9FC"
BLOCK_FILL = "1F4E79"

thin = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
thick_blue = Border(
    left=Side(style="medium", color=BLUE),
    right=Side(style="medium", color=BLUE),
    top=Side(style="medium", color=BLUE),
    bottom=Side(style="medium", color=BLUE),
)


def fix_workbook(src: Path, dst: Path) -> None:
    with zipfile.ZipFile(src, "r") as zin, zipfile.ZipFile(
        dst, "w", compression=zipfile.ZIP_DEFLATED
    ) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/styles.xml":
                text = data.decode("utf-8")
                text = text.replace('style="solid"', 'style="thin"')
                text = re.sub(r'\s*style="none"', "", text)
                text = re.sub(r'rgb="none"', 'rgb="FF000000"', text)
                data = text.encode("utf-8")
            zout.writestr(item, data)


def style_range(ws, cells, **kwargs):
    for coord in cells:
        cell = ws[coord] if isinstance(coord, str) else ws.cell(*coord)
        for k, v in kwargs.items():
            setattr(cell, k, v)


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def font(bold=False, size=11, color="000000", name="Calibri"):
    return Font(name=name, bold=bold, size=size, color=color)


def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height


def merge_write(ws, range_str, value, **style):
    ws.merge_cells(range_str)
    cell = ws[range_str.split(":")[0]]
    cell.value = value
    for k, v in style.items():
        setattr(cell, k, v)
    return cell


def apply_box(ws, start_row, end_row, start_col, end_col, border=thin):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(r, c).border = border


def clear_sheet(ws):
    # unmerge all
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    # clear values and styles in used area
    for row in ws.iter_rows(min_row=1, max_row=max(ws.max_row, 1), max_col=max(ws.max_column, 1)):
        for cell in row:
            cell.value = None
            cell.fill = PatternFill()
            cell.font = Font()
            cell.border = Border()
            cell.alignment = Alignment()
            cell.number_format = "General"


def build_latest_sheet(ws):
    clear_sheet(ws)

    # Column widths
    widths = {
        "A": 6,
        "B": 42,
        "C": 28,
        "D": 18,
        "E": 16,
        "F": 36,
        "G": 18,
        "H": 14,
        "I": 14,
        "J": 22,
        "K": 28,
        "L": 3,
        "M": 14,
        "N": 12,
        "O": 28,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ========== TITLE ==========
    merge_write(
        ws,
        "A1:G1",
        "ВСТРЕЧА ЦП — ЦКС — ЦСВ — СМАК  ·  Протокол 27.08.2026",
        font=font(True, 16, WHITE),
        fill=fill(BLUE),
        alignment=align("center", "center"),
    )
    set_row_height(ws, 1, 32)

    merge_write(
        ws,
        "A2:G2",
        "Проект ЭПД  ·  еженедельный статус",
        font=font(False, 11, BLUE),
        fill=fill(BLUE_PALE),
        alignment=align("center", "center"),
    )
    set_row_height(ws, 2, 20)

    # ========== PLAN BLOCK ==========
    # Left: plan summary + departments
    # Right: period comparison

    merge_write(
        ws,
        "A4:G4",
        "ПЛАН ПО ЭПД",
        font=font(True, 13, WHITE),
        fill=fill(BLUE_MID),
        alignment=align("left", "center"),
    )
    set_row_height(ws, 4, 24)

    # Summary headers
    headers = [
        ("A5", "План"),
        ("C5", "Текущее значение"),
        ("E5", "Итого до плана"),
        ("G5", "% выполнения"),
    ]
    for coord, text in headers:
        cell = ws[coord]
        cell.value = text
        cell.font = font(True, 10, GRAY)
        cell.fill = fill(BLUE_LIGHT)
        cell.alignment = align("center", "center")
        cell.border = thin

    ws.merge_cells("A5:B5")
    ws.merge_cells("C5:D5")
    ws.merge_cells("E5:F5")

    # Values row
    # Plan = 500 in B6 (named conceptually)
    # Current = SUM of department facts N11:N13 (see department table)
    # Remaining = plan - current
    # % = current / plan

    ws["A6"].value = 500
    ws["A6"].font = font(True, 18, BLUE)
    ws["A6"].fill = fill(BLUE_PALE)
    ws["A6"].alignment = align("center", "center")
    ws["A6"].border = thin
    ws["A6"].number_format = '#,##0" клиент."'
    ws.merge_cells("A6:B6")
    ws["B6"].border = thin
    ws["B6"].fill = fill(BLUE_PALE)

    # Current value linked to department facts (N11+N12+N13)
    # Also show note that it's sum of «Факт» by departments
    ws["C6"].value = "=N11+N12+N13"
    ws["C6"].font = font(True, 18, GREEN)
    ws["C6"].fill = fill(GREEN_LIGHT)
    ws["C6"].alignment = align("center", "center")
    ws["C6"].border = thin
    ws["C6"].number_format = '#,##0" клиент."'
    ws.merge_cells("C6:D6")
    ws["D6"].border = thin
    ws["D6"].fill = fill(GREEN_LIGHT)

    ws["E6"].value = "=A6-C6"
    ws["E6"].font = font(True, 18, ORANGE)
    ws["E6"].fill = fill(ORANGE_LIGHT)
    ws["E6"].alignment = align("center", "center")
    ws["E6"].border = thin
    ws["E6"].number_format = '#,##0" клиент."'
    ws.merge_cells("E6:F6")
    ws["F6"].border = thin
    ws["F6"].fill = fill(ORANGE_LIGHT)

    ws["G6"].value = '=IF(A6=0,"—",C6/A6)'
    ws["G6"].font = font(True, 18, BLUE)
    ws["G6"].fill = fill(YELLOW)
    ws["G6"].alignment = align("center", "center")
    ws["G6"].border = thin
    ws["G6"].number_format = "0.0%"

    set_row_height(ws, 5, 18)
    set_row_height(ws, 6, 36)

    merge_write(
        ws,
        "A7:G7",
        "«Текущее значение» = сумма ячеек «Факт» по отделам (ниже). Меняйте факты — обновятся итог, остаток и % выполнения.",
        font=font(False, 9, GRAY),
        fill=fill(GRAY_LIGHT),
        alignment=align("left", "center"),
    )
    set_row_height(ws, 7, 18)

    # Department plans table (placed to the RIGHT of summary for visibility — user asked "по отделам" under plan)
    merge_write(
        ws,
        "A9:G9",
        "План на МПП в месяц — по отделам",
        font=font(True, 11, WHITE),
        fill=fill(GREEN),
        alignment=align("left", "center"),
    )

    dept_headers = ["Отдел / группа", "План на МПП в мес", "Факт", "% выполнения плана"]
    # Use columns A-B, C, D, E-F for department table on left? 
    # User also wants facts linked - I'll put editable fact cells in a clear place.
    # Better layout: department table in A-G rows 10-13, AND mirror fact values also in N11:N13
    # Actually simpler: put department table with facts in columns that C6 references.
    # Put department table starting at column I? User said "табличку сбоку" for period comparison.
    # So: departments under plan (left), period comparison on the right (I-K).

    # Department table headers row 10
    ws["A10"].value = "Отдел / группа"
    ws["C10"].value = "План на МПП в мес"
    ws["E10"].value = "Факт"
    ws["F10"].value = "% выполнения плана"
    ws.merge_cells("A10:B10")
    ws.merge_cells("C10:D10")
    ws.merge_cells("F10:G10")
    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        ws[f"{col}10"].font = font(True, 10, WHITE)
        ws[f"{col}10"].fill = fill(GREEN)
        ws[f"{col}10"].alignment = align("center", "center")
        ws[f"{col}10"].border = thin

    departments = [
        (11, "Группа «Новые деньги»", 75, 90),   # facts sum to 150 with row12+13
        (12, "Группа продуктового запуска", 25, 40),
        (13, "ЦКС", None, 20),
    ]
    # 90+40+20 = 150 ✓

    for row, name, plan, fact in departments:
        ws.merge_cells(f"A{row}:B{row}")
        ws[f"A{row}"].value = name
        ws[f"A{row}"].font = font(True, 10)
        ws[f"A{row}"].alignment = align("left", "center")
        ws[f"A{row}"].fill = fill(GREEN_LIGHT)
        ws[f"B{row}"].fill = fill(GREEN_LIGHT)
        ws[f"B{row}"].border = thin
        ws[f"A{row}"].border = thin

        ws.merge_cells(f"C{row}:D{row}")
        plan_fill = fill(GRAY_LIGHT) if plan is None else fill(WHITE)
        if plan is None:
            ws[f"C{row}"].value = None  # empty for CKS as requested
        else:
            ws[f"C{row}"].value = plan
        ws[f"C{row}"].fill = plan_fill
        ws[f"C{row}"].font = font(True, 12, BLUE)
        ws[f"C{row}"].alignment = align("center", "center")
        ws[f"C{row}"].border = thin
        ws[f"C{row}"].number_format = "0"
        ws[f"D{row}"].border = thin
        ws[f"D{row}"].fill = plan_fill

        # Fact — editable, yellow highlight
        ws[f"E{row}"].value = fact
        ws[f"E{row}"].font = font(True, 12, GREEN)
        ws[f"E{row}"].fill = fill(YELLOW)
        ws[f"E{row}"].alignment = align("center", "center")
        ws[f"E{row}"].border = thin
        ws[f"E{row}"].number_format = "0"

        # % = fact/plan
        ws.merge_cells(f"F{row}:G{row}")
        ws[f"F{row}"].value = f'=IF(OR(C{row}="",C{row}=0),"—",E{row}/C{row})'
        ws[f"F{row}"].font = font(True, 11)
        ws[f"F{row}"].alignment = align("center", "center")
        ws[f"F{row}"].border = thin
        ws[f"F{row}"].number_format = "0.0%"
        ws[f"G{row}"].border = thin
        set_row_height(ws, row, 22)

    # Hidden helper cells N11:N13 = references to E11:E13 (so C6 formula is stable/documented)
    # Actually C6 already can use E11+E12+E13 directly — update C6
    ws["C6"].value = "=E11+E12+E13"

    # Also keep N11:N13 as named helpers for period table clarity (optional mirror)
    for r in (11, 12, 13):
        ws[f"N{r}"].value = f"=E{r}"

    # ========== SIDE TABLE: period comparison ==========
    merge_write(
        ws,
        "I4:K4",
        "Соотношение к прошлому периоду",
        font=font(True, 11, WHITE),
        fill=fill(ORANGE),
        alignment=align("center", "center"),
    )
    set_row_height(ws, 4, 24)

    for col, text in [("I5", "Дата"), ("J5", "Клиенты"), ("K5", "Изменение")]:
        ws[col].value = text
        ws[col].font = font(True, 10, WHITE)
        ws[col].fill = fill(ORANGE)
        ws[col].alignment = align("center", "center")
        ws[col].border = thin

    # Period rows — first period baseline, then examples + empty rows for future
    # Row 6: 27.08.2026 — current (=C6)
    # Row 7: template 03.09 — example empty for user
    # Row 8: template 10.09
    # Plus 2 more empty rows

    periods = [
        (6, "27.08.2026", "=C6", None),  # linked to current
        (7, "03.09.2026", None, None),   # user fills
        (8, "10.09.2026", None, None),
        (9, "", None, None),
        (10, "", None, None),
        (11, "", None, None),
        (12, "", None, None),
        (13, "", None, None),
    ]

    # For demo of the formula: put example values in a note area and also
    # seed 03.09 / 10.09 with the user's example (200 → 250) as illustration,
    # while 27.08 stays linked to current (150).
    # User example: 03.09=200, 10.09=250 → +25%
    # I'll put 27.08 linked, then leave 03.09 and 10.09 as editable examples with 200/250
    # so the % formula is visible immediately.

    ws["I6"].value = "27.08.2026"
    ws["J6"].value = "=C6"  # 150 via facts
    ws["K6"].value = "—"

    ws["I7"].value = "03.09.2026"
    ws["J7"].value = 200  # example / editable
    ws["K7"].value = '=IF(OR(J6="",J6=0,J7=""),"—",IF(J7>=J6,"увеличилось на "&TEXT((J7-J6)/J6,"0.0%"),"уменьшилось на "&TEXT((J6-J7)/J6,"0.0%")))'

    ws["I8"].value = "10.09.2026"
    ws["J8"].value = 250  # example / editable
    ws["K8"].value = '=IF(OR(J7="",J7=0,J8=""),"—",IF(J8>=J7,"увеличилось на "&TEXT((J8-J7)/J7,"0.0%"),"уменьшилось на "&TEXT((J7-J8)/J7,"0.0%")))'

    for r in range(9, 14):
        ws[f"K{r}"].value = (
            f'=IF(OR(J{r-1}="",J{r-1}=0,J{r}=""),"—",'
            f'IF(J{r}>=J{r-1},"увеличилось на "&TEXT((J{r}-J{r-1})/J{r-1},"0.0%"),'
            f'"уменьшилось на "&TEXT((J{r-1}-J{r})/J{r-1},"0.0%")))'
        )

    for r in range(6, 14):
        for col in ["I", "J", "K"]:
            ws[f"{col}{r}"].border = thin
            ws[f"{col}{r}"].alignment = align("center", "center")
            ws[f"{col}{r}"].font = font(False, 10)
        ws[f"J{r}"].fill = fill(YELLOW)
        ws[f"J{r}"].number_format = "0"
        ws[f"K{r}"].fill = fill(ORANGE_LIGHT)
        ws[f"K{r}"].alignment = align("left", "center")
        set_row_height(ws, r, 22)

    ws["J6"].fill = fill(GREEN_LIGHT)  # auto-linked
    ws["I6"].font = font(True, 10)

    merge_write(
        ws,
        "I14:K14",
        "Жёлтые ячейки «Клиенты» — вводите вручную на новые даты. Строка 27.08 подтягивает текущее значение из плана.",
        font=font(False, 8, GRAY),
        fill=fill(GRAY_LIGHT),
        alignment=align("left", "center"),
    )
    set_row_height(ws, 14, 28)

    # ========== AD CAMPAIGN (compact) ==========
    r = 16
    merge_write(
        ws,
        f"A{r}:G{r}",
        "Показатели рекламной кампании",
        font=font(True, 12, WHITE),
        fill=fill(BLUE),
        alignment=align("left", "center"),
    )

    r = 17
    for col, text in [
        ("A", "№"),
        ("B", "Название РК"),
        ("C", "Лидер"),
        ("D", "Кол-во ЛИДов"),
        ("E", "Стоимость ЛИДа"),
        ("F", "Сделок"),
        ("G", "Оплат"),
    ]:
        ws[f"{col}{r}"].value = text
        ws[f"{col}{r}"].font = font(True, 10, WHITE)
        ws[f"{col}{r}"].fill = fill(BLUE_MID)
        ws[f"{col}{r}"].alignment = align("center", "center")
        ws[f"{col}{r}"].border = thin

    r = 18
    ws["A18"] = 1
    ws["B18"] = "ЭПД"
    ws["C18"] = "Савинская"
    ws["D18"] = (
        "ЭПД Квиз — 9 шт; CRM-форма «Внедрение и поддержка 1С:ЭПД» — 17 шт; "
        "Заявка с сайта — 6 шт"
    )
    ws["E18"] = 1400
    ws["F18"] = (
        "Квиз — 3 сделки (2 контроль оплаты); CRM-форма — 5 сделок "
        "(1 контроль оплаты на 18 100); Заявка с сайта — 5 сделок (2 оплаты, 3 контроль)"
    )
    ws["G18"] = "2 шт / 10 000 ₽"
    for col in "ABCDEFG":
        ws[f"{col}18"].border = thin
        ws[f"{col}18"].alignment = align("left", "center")
        ws[f"{col}18"].fill = fill(BLUE_PALE)
    ws["E18"].number_format = '#,##0" ₽"'
    ws["E18"].alignment = align("center", "center")
    set_row_height(ws, 18, 48)

    # ========== 1. MONITORING ==========
    r = 20
    merge_write(
        ws,
        f"A{r}:G{r}",
        "1. Мониторинг предыдущих решений",
        font=font(True, 12, WHITE),
        fill=fill(BLUE),
        alignment=align("left", "center"),
    )

    r = 21
    for col, text in [
        ("A", "№"),
        ("B", "Задача"),
        ("C", "Лидер"),
        ("D", "Срок"),
        ("E", "Статус"),
        ("F", "Комментарии / выводы"),
        ("G", ""),
    ]:
        ws[f"{col}{r}"].value = text
        ws[f"{col}{r}"].font = font(True, 10, WHITE)
        ws[f"{col}{r}"].fill = fill(BLUE_MID)
        ws[f"{col}{r}"].alignment = align("center", "center")
        ws[f"{col}{r}"].border = thin
    ws.merge_cells("F21:G21")

    monitoring = [
        (1, "Включить поле «время взятия в работу» заявки с РК", "Савинская", "до 27.08", "выполнено", ""),
        (2, "Обсудить разделение стоимости РК между ЦП и ЦКС на очной встрече 27.08", "Антонов В., Дербенева К. (все)", "до 27.08", "в работе", ""),
        (3, "Текущая РК действует до середины сентября. Продлеваем? Согласование бюджета", "Все участники", "до 27.08", "в работе", ""),
        (4, "Собрать ОС от МПП по ЭПД — проблемы, возражения, потребности", "Савинская / Блохина / Степанова", "до 24.08", "выполнено", "Документ в беседе «ЭПД стратегия»"),
        (5, "Собрать портрет клиента, который купил ЭПД", "Корнева / Степанова", "до 27.08", "в работе", "Параметры портрета — после встречи"),
        (6, "Добавить Корневу и Степанову во встречу с Калугой", "Блохина", "до 24.08", "выполнено", ""),
        (7, "Построить финмодель направления ЭПД (доходы, затраты, конверсии, сценарии)", "Все участники", "после стратег-сессии", "в работе", ""),
    ]

    for i, (num, task, leader, deadline, status, comment) in enumerate(monitoring):
        row = 22 + i
        ws[f"A{row}"] = num
        ws[f"B{row}"] = task
        ws[f"C{row}"] = leader
        ws[f"D{row}"] = deadline
        ws[f"E{row}"] = status
        ws[f"F{row}"] = comment
        ws.merge_cells(f"F{row}:G{row}")
        for col in "ABCDEFG":
            ws[f"{col}{row}"].border = thin
            ws[f"{col}{row}"].alignment = align("left", "center")
            ws[f"{col}{row}"].font = font(False, 10)
        ws[f"A{row}"].alignment = align("center", "center")
        ws[f"E{row}"].alignment = align("center", "center")
        ws[f"E{row}"].font = font(True, 10)
        if "выполн" in status.lower():
            ws[f"E{row}"].fill = fill(GREEN_LIGHT)
            ws[f"E{row}"].font = font(True, 10, GREEN)
        else:
            ws[f"E{row}"].fill = fill(ORANGE_LIGHT)
            ws[f"E{row}"].font = font(True, 10, ORANGE)
        set_row_height(ws, row, 32)

    # ========== 2. CURRENT DECISIONS / TASKS ==========
    r = 30
    merge_write(
        ws,
        f"A{r}:G{r}",
        "2. Решения текущей встречи  ·  поставленные задачи",
        font=font(True, 12, WHITE),
        fill=fill(ORANGE),
        alignment=align("left", "center"),
    )

    r = 31
    for col, text in [
        ("A", "№"),
        ("B", "Задача"),
        ("C", "Лидер"),
        ("D", "Срок"),
        ("E", "Статус"),
        ("F", "Комментарии / выводы"),
        ("G", ""),
    ]:
        ws[f"{col}{r}"].value = text
        ws[f"{col}{r}"].font = font(True, 10, WHITE)
        ws[f"{col}{r}"].fill = fill(ORANGE)
        ws[f"{col}{r}"].alignment = align("center", "center")
        ws[f"{col}{r}"].border = thin
    ws.merge_cells("F31:G31")

    # Extracted & structured tasks from discussion
    decisions = [
        (
            1,
            "Разобраться с показателями охвата сервисом 1С-ЭПД из таблицы А. Дворяк (9 000 ед.): что считается охватом — подключение или предоплаченные пакеты?",
            "Корнева, Степанова",
            "до след. встречи",
            "в работе",
            "Цифры расходятся с Калугой; проблемы с закреплением",
        ),
        (
            2,
            "Уточнить у Суворовой статус Элиттрейд; предложить акцию 3 мес. Доки для захода к контрагентам",
            "Блохина (Лиза)",
            "до след. встречи",
            "в работе",
            "Крупные производители «диктуют» правила → возможность почкования базы",
        ),
        (
            3,
            "Подключить Антона Гуркова к процессу изучения привязки клиента (Доки.Логистика / ЦКС)",
            "Гурков А.",
            "в работе",
            "взял на себя",
            "Задача Оксаны Ремез — прописать подробнее после пересмотра встречи",
        ),
        (
            4,
            "Выяснить у Калуги: можно ли отследить активность клиента в сервисе Доки",
            "Дивакова М.",
            "до след. встречи",
            "в работе",
            "",
        ),
        (
            5,
            "Уточнить у Никитченко: считается ли Доки в охват сервисами ЭПД",
            "Дворяк (Саша)",
            "до след. встречи",
            "в работе",
            "Саша спросит",
        ),
        (
            6,
            "Подготовить со СМАК материалы для прогрева клиента на период демо-доступа",
            "СМАК + ответственный",
            "к 01.10",
            "в работе",
            "С 01.10 демо = 14 дней; до 01.10 — 3 месяца бесплатно",
        ),
        (
            7,
            "Подготовить калькулятор пополнения титулов / помощи менеджеру ЦКС при продлении",
            "Вика + Маша",
            "до след. встречи",
            "в работе",
            "Обсуждали накануне лектория с 1С",
        ),
        (
            8,
            "Подготовить текст рассылки: «проверьте своего контрагента и впишите в тандем» — показать сроки и текст",
            "Блохина (Лиза)",
            "до след. встречи",
            "в работе",
            "",
        ),
        (
            9,
            "Обзвонить крупных клиентов (от 2 000 титулов / генераторы трафика) и предложить ЭПД ТАНДЕМ",
            "ЦП / ЦКС",
            "сентябрь",
            "в работе",
            "Начать с крупных; корп. сегмент продавал ЭПД даже в малом объёме",
        ),
        (
            10,
            "Обзвон подключённых клиентов: оценка удовлетворённости + проекты ЦАС (Доки)",
            "Дворяк / Лазарчук",
            "сентябрь",
            "в работе",
            "Задача по клиентам ЦАС (Доки)",
        ),
        (
            11,
            "Взять неуспешные сделки ЭПД по корп. сегменту и предложить Доки в первую неделю сентября",
            "Юлиана",
            "1-я неделя сентября",
            "в работе",
            "Клиенты сбытового офиса уходят на СБИС — разобрать причины",
        ),
        (
            12,
            "Проверить в базе: подключён ли клиент к каналу ЭПД (МАКС / ТГ / ВК); оценить как альт. канал привлечения",
            "ЦП / ЦКС",
            "до след. встречи",
            "в работе",
            "Связать со стратегией и целями",
        ),
        (
            13,
            "Зафиксировать целевую структуру плана 500 клиентов до конца 2026: 50 крупных / 300 средних / 150 мелких",
            "Все участники",
            "принято",
            "принято",
            "Приоритет — количество клиентов, не сумма выручки на старте",
        ),
    ]

    for i, (num, task, leader, deadline, status, comment) in enumerate(decisions):
        row = 32 + i
        ws[f"A{row}"] = num
        ws[f"B{row}"] = task
        ws[f"C{row}"] = leader
        ws[f"D{row}"] = deadline
        ws[f"E{row}"] = status
        ws[f"F{row}"] = comment
        ws.merge_cells(f"F{row}:G{row}")
        for col in "ABCDEFG":
            ws[f"{col}{row}"].border = thin
            ws[f"{col}{row}"].alignment = align("left", "center")
            ws[f"{col}{row}"].font = font(False, 10)
            ws[f"{col}{row}"].fill = fill(TASK_FILL)
        ws[f"A{row}"].alignment = align("center", "center")
        ws[f"A{row}"].font = font(True, 10, ORANGE)
        ws[f"E{row}"].alignment = align("center", "center")
        ws[f"E{row}"].font = font(True, 10)
        st = status.lower()
        if "принят" in st or "выполн" in st:
            ws[f"E{row}"].fill = fill(GREEN_LIGHT)
            ws[f"E{row}"].font = font(True, 10, GREEN)
        elif "взял" in st:
            ws[f"E{row}"].fill = fill(BLUE_LIGHT)
            ws[f"E{row}"].font = font(True, 10, BLUE)
        else:
            ws[f"E{row}"].fill = fill(ORANGE_LIGHT)
            ws[f"E{row}"].font = font(True, 10, ORANGE)
        set_row_height(ws, row, 42)

    last_decision_row = 32 + len(decisions) - 1  # 44

    # ========== 3. DISCUSSION THEMES ==========
    r = last_decision_row + 2  # 46
    merge_write(
        ws,
        f"A{r}:G{r}",
        "3. Обсуждаемые темы и выводы  ·  блоки и тезисы",
        font=font(True, 12, WHITE),
        fill=fill(BLUE),
        alignment=align("left", "center"),
    )
    section3_start = r

    r += 1
    ws[f"A{r}"] = "Блок"
    ws.merge_cells(f"B{r}:C{r}")
    ws[f"B{r}"] = "Тема / тезис"
    ws.merge_cells(f"D{r}:G{r}")
    ws[f"D{r}"] = "Выводы и решения"
    for col in "ABCDEFG":
        ws[f"{col}{r}"].font = font(True, 10, WHITE)
        ws[f"{col}{r}"].fill = fill(BLUE_MID)
        ws[f"{col}{r}"].alignment = align("center", "center")
        ws[f"{col}{r}"].border = thin
    header_row = r

    # Theme blocks with color coding
    blocks = [
        {
            "name": "Акция и ЦСВ",
            "color": "5B9BD5",
            "items": [
                (
                    "Трудозатраты ЦСВ на акцию по ЭПД",
                    "Акцию продлили до 30.09. Уточнить адресата вопроса у Оксаны.",
                ),
            ],
        },
        {
            "name": "Каналы\nМАКС / ТГ / ВК",
            "color": "70AD47",
            "items": [
                (
                    "Можем ли прорабатывать клиентов канала ЭПД силами ЦП и ЦКС? Альтернативный канал привлечения.",
                    "Нужна проверка базы: подключён клиент к каналу или нет. Идея — просить менеджеров подключать клиентов; вопрос мотивации клиента открыт.",
                ),
            ],
        },
        {
            "name": "Стратегия\nи сегментация",
            "color": "ED7D31",
            "items": [
                (
                    "Заход в Элиттрейд (таблица Дворяк) → контрагенты. Сегментировать охваченных на крупных / средних / мелких.",
                    "Производители (крупные) задают правила для перевозчиков → те для получателей. Лиза уточнит у Суворовой статус Элиттрейд + акция 3 мес. Доки.",
                ),
                (
                    "Признак «крупный» = генератор трафика (привлечение контрагентов).",
                    "Согласовали целевое соотношение базы под план 500.",
                ),
                (
                    "Приоритет — количество привлечённых клиентов, не сумма заработка. Выстроить цепочку «закрепления» и почкования с одного пакета.",
                    "ПЛАН: 500 клиентов до конца 2026 → 50 крупных + 300 средних + 150 мелких.",
                ),
            ],
        },
        {
            "name": "Доки.\nЛогистика\nи охват",
            "color": "9E480E",
            "items": [
                (
                    "Задача по Доки.Логистика (Оксана Ремез) — детализировать после пересмотра встречи.",
                    "Подключить Антона Гуркова к изучению привязки клиента. Антон взял на себя.",
                ),
                (
                    "Можем ли отследить активность клиента в Доки? (вопрос Калуге)",
                    "Маша Дивакова выяснит.",
                ),
                (
                    "Считается ли Доки в охват сервисами ЭПД?",
                    "Вопрос Никитченко — Саша спросит. Цифры расходятся с Калугой; проблемы с закреплением.",
                ),
                (
                    "Все проданные ЭПД: центр компетенции изучает — всё ли продано, подключились ли.",
                    "Трафик не видим → не квалифицируем активность. 1С обещают данные в сентябре.",
                ),
            ],
        },
        {
            "name": "Демо\nи прогрев",
            "color": "7030A0",
            "items": [
                (
                    "Со СМАК подготовить материалы прогрева на период демо-доступа.",
                    "С 01.10 — 14 дней демо; до 01.10 — 3 месяца бесплатно.",
                ),
            ],
        },
        {
            "name": "Сопровождение\nМПП / ЦКС",
            "color": "C00000",
            "items": [
                (
                    "План МПП: платные / бесплатные клиенты; «выхаживать» может отдельный человек (идея — Гурков).",
                    "Если есть ИТС — ЦКС смотрит динамику расхода ЭПД/Доки. Трафика пока нет; можно писать в Калугу с перечнем ИНН. Риск: кто выхаживает — текущие МПП или новый человек?",
                ),
                (
                    "Клиенту продают мин. пакет («пока не понятно»). Кто ведёт дальше? Когда собирать ОС перед продлением?",
                    "По логике — ЦКС. Нужен калькулятор — Вика с Машей обсуждали накануне лектория.",
                ),
                (
                    "ЦКС категоризирует клиента по титулам / УАТ и вовремя пополняет баланс.",
                    "Ожидание зафиксировано как рабочий процесс.",
                ),
            ],
        },
        {
            "name": "Тандем\nи обзвоны",
            "color": "00B0F0",
            "items": [
                (
                    "Рассылка: «проверьте контрагента и впишите в тандем».",
                    "Лиза — когда и какой текст.",
                ),
                (
                    "Начать с крупных: прозвонить и предложить подключить контрагента в ЭПД ТАНДЕМ.",
                    "Крупные = от 2 000 титулов. Корп. клиенты покупали ЭПД даже в малом объёме.",
                ),
                (
                    "Обзвон подключённых: оценка удовлетворённости + проекты ЦАС (Доки).",
                    "Задача на Сашу Дворяк / Лазарчук — клиенты ЦАС.",
                ),
            ],
        },
        {
            "name": "Корп.\nсегмент",
            "color": "548235",
            "items": [
                (
                    "Почему клиенты сбытового офиса не переходят в ЭПД / уходят на СБИС?",
                    "Взять неуспешные сделки корп. сегмента и предложить Доки в 1-ю неделю сентября (Юлиана).",
                ),
                (
                    "Юля: предлагать всем бесплатно Доки (дубль) на 3 мес.; KPI менеджерам на сентябрь.",
                    "Зафиксировано как предложение.",
                ),
            ],
        },
    ]

    row = header_row + 1
    block_fills = [
        "DDEBF7",
        "E2EFDA",
        "FCE4D6",
        "F8CBAD",
        "E2D5F1",
        "F8D7DA",
        "D6F0FA",
        "E2EFD9",
    ]

    for bi, block in enumerate(blocks):
        bg = block_fills[bi % len(block_fills)]
        start = row
        for topic, conclusion in block["items"]:
            ws[f"A{row}"] = block["name"] if row == start else ""
            ws.merge_cells(f"B{row}:C{row}")
            ws[f"B{row}"] = topic
            ws.merge_cells(f"D{row}:G{row}")
            ws[f"D{row}"] = conclusion
            for col in "ABCDEFG":
                ws[f"{col}{row}"].border = thin
                ws[f"{col}{row}"].fill = fill(bg)
                ws[f"{col}{row}"].alignment = align("left", "center")
                ws[f"{col}{row}"].font = font(False, 10)
            ws[f"A{row}"].font = font(True, 9, WHITE)
            ws[f"A{row}"].fill = fill(block["color"])
            ws[f"A{row}"].alignment = align("center", "center")
            ws[f"B{row}"].font = font(True, 10)
            ws[f"D{row}"].fill = fill(THESIS_FILL) if bi % 2 == 0 else fill(bg)
            # highlight conclusions that contain tasks / plan
            text_l = conclusion.lower()
            if any(k in text_l for k in ("план:", "взял", "выяснит", "спросит", "лиза", "юлиана", "задача")):
                ws[f"D{row}"].fill = fill(TASK_FILL)
                ws[f"D{row}"].font = font(False, 10)
            set_row_height(ws, row, 48)
            row += 1
        end = row - 1
        if end > start:
            ws.merge_cells(f"A{start}:A{end}")
            ws[f"A{start}"].alignment = align("center", "center")

    legend_row = row + 1
    merge_write(
        ws,
        f"A{legend_row}:G{legend_row}",
        "Легенда:  оранжевый блок «Решения» = поставленные задачи  ·  жёлтые ячейки = ввод факта  ·  зелёный статус = выполнено/принято  ·  блоки слева = темы обсуждения",
        font=font(False, 9, GRAY),
        fill=fill(GRAY_LIGHT),
        alignment=align("left", "center"),
    )

    # Freeze panes below title
    ws.freeze_panes = "A4"
    ws.print_title_rows = "1:2"
    ws.page_setup.orientation = "landscape"

    # Data validation for status columns
    dv = DataValidation(
        type="list",
        formula1='"в работе,выполнено,принято,взял на себя,отложено"',
        allow_blank=True,
    )
    dv.error = "Выберите статус из списка"
    dv.errorTitle = "Статус"
    ws.add_data_validation(dv)
    dv.add("E22:E28")
    dv.add("E32:E44")


def main():
    src = ROOT / SRC_NAME
    if not src.exists():
        raise SystemExit(f"Source not found: {src}")

    fix_workbook(src, FIXED_TMP)
    wb = load_workbook(FIXED_TMP)

    # Rename / rebuild latest sheet
    if "27.08.2026" in wb.sheetnames:
        ws = wb["27.08.2026"]
    else:
        ws = wb.active

    build_latest_sheet(ws)

    # Add a clean template sheet for next meeting at the front-ish
    if "Шаблон (новая встреча)" in wb.sheetnames:
        del wb["Шаблон (новая встреча)"]
    tpl = wb.create_sheet("Шаблон (новая встреча)", 0)
    build_template(tpl)

    # Move formatted latest to position 1 (after template)
    wb.move_sheet(ws, offset=1 - wb.sheetnames.index(ws.title))

    wb.save(OUT_PATH)
    print(f"Saved: {OUT_PATH}")

    # Verify formulas
    wb2 = load_workbook(OUT_PATH)
    ws2 = wb2["27.08.2026"]
    print("C6 (current):", ws2["C6"].value)
    print("E6 (remaining):", ws2["E6"].value)
    print("G6 (%):", ws2["G6"].value)
    print("E11-13 facts:", ws2["E11"].value, ws2["E12"].value, ws2["E13"].value)
    print("K7 change:", ws2["K7"].value)
    print("K8 change:", ws2["K8"].value)
    print("Sheets:", wb2.sheetnames)


def build_template(ws):
    """Lightweight template for copying to a new meeting date."""
    clear_sheet(ws)
    for col, w in {
        "A": 6, "B": 42, "C": 28, "D": 18, "E": 16, "F": 36, "G": 18,
        "H": 14, "I": 14, "J": 12, "K": 28,
    }.items():
        ws.column_dimensions[col].width = w

    merge_write(
        ws, "A1:G1",
        "ВСТРЕЧА ЦП — ЦКС — ЦСВ — СМАК  ·  Протокол ДД.ММ.ГГГГ",
        font=font(True, 16, WHITE), fill=fill(BLUE), alignment=align("center", "center"),
    )
    set_row_height(ws, 1, 32)
    merge_write(
        ws, "A2:G2",
        "Скопируйте лист и переименуйте в дату встречи. Обновите факты и таблицу периодов.",
        font=font(False, 10, BLUE), fill=fill(BLUE_PALE), alignment=align("center", "center"),
    )

    merge_write(ws, "A4:G4", "ПЛАН ПО ЭПД", font=font(True, 13, WHITE), fill=fill(BLUE_MID), alignment=align("left", "center"))
    for coord, text in [("A5", "План"), ("C5", "Текущее значение"), ("E5", "Итого до плана"), ("G5", "% выполнения")]:
        ws[coord] = text
        ws[coord].font = font(True, 10, GRAY)
        ws[coord].fill = fill(BLUE_LIGHT)
        ws[coord].alignment = align("center", "center")
        ws[coord].border = thin
    ws.merge_cells("A5:B5"); ws.merge_cells("C5:D5"); ws.merge_cells("E5:F5")

    ws["A6"] = 500
    ws.merge_cells("A6:B6")
    ws["C6"] = "=E11+E12+E13"
    ws.merge_cells("C6:D6")
    ws["E6"] = "=A6-C6"
    ws.merge_cells("E6:F6")
    ws["G6"] = '=IF(A6=0,"—",C6/A6)'
    ws["G6"].number_format = "0.0%"
    for coord in ["A6", "C6", "E6", "G6"]:
        ws[coord].font = font(True, 16, BLUE)
        ws[coord].alignment = align("center", "center")
        ws[coord].border = thin
        ws[coord].fill = fill(BLUE_PALE)
    ws["C6"].fill = fill(GREEN_LIGHT)
    ws["E6"].fill = fill(ORANGE_LIGHT)
    ws["G6"].fill = fill(YELLOW)
    ws["A6"].number_format = "0"
    ws["C6"].number_format = "0"
    ws["E6"].number_format = "0"

    merge_write(ws, "A9:G9", "План на МПП в месяц — по отделам", font=font(True, 11, WHITE), fill=fill(GREEN), alignment=align("left", "center"))
    ws["A10"] = "Отдел / группа"; ws.merge_cells("A10:B10")
    ws["C10"] = "План на МПП в мес"; ws.merge_cells("C10:D10")
    ws["E10"] = "Факт"
    ws["F10"] = "% выполнения плана"; ws.merge_cells("F10:G10")
    for col in "ABCDEFG":
        ws[f"{col}10"].font = font(True, 10, WHITE)
        ws[f"{col}10"].fill = fill(GREEN)
        ws[f"{col}10"].border = thin
        ws[f"{col}10"].alignment = align("center", "center")

    for row, name, plan in [(11, "Группа «Новые деньги»", 75), (12, "Группа продуктового запуска", 25), (13, "ЦКС", None)]:
        ws.merge_cells(f"A{row}:B{row}")
        ws[f"A{row}"] = name
        ws.merge_cells(f"C{row}:D{row}")
        ws[f"C{row}"] = plan
        ws[f"E{row}"] = None
        ws[f"E{row}"].fill = fill(YELLOW)
        ws.merge_cells(f"F{row}:G{row}")
        ws[f"F{row}"] = f'=IF(OR(C{row}="",C{row}=0),"—",E{row}/C{row})'
        ws[f"F{row}"].number_format = "0.0%"
        for col in "ABCDEFG":
            ws[f"{col}{row}"].border = thin
            ws[f"{col}{row}"].alignment = align("center", "center")

    merge_write(ws, "I4:K4", "Соотношение к прошлому периоду", font=font(True, 11, WHITE), fill=fill(ORANGE), alignment=align("center", "center"))
    for col, t in [("I5", "Дата"), ("J5", "Клиенты"), ("K5", "Изменение")]:
        ws[col] = t
        ws[col].font = font(True, 10, WHITE)
        ws[col].fill = fill(ORANGE)
        ws[col].border = thin
        ws[col].alignment = align("center", "center")
    ws["I6"] = "← дата"
    ws["J6"] = "=C6"
    ws["K6"] = "—"
    for r in range(7, 12):
        ws[f"K{r}"] = (
            f'=IF(OR(J{r-1}="",J{r-1}=0,J{r}=""),"—",'
            f'IF(J{r}>=J{r-1},"увеличилось на "&TEXT((J{r}-J{r-1})/J{r-1},"0.0%"),'
            f'"уменьшилось на "&TEXT((J{r-1}-J{r})/J{r-1},"0.0%")))'
        )
        ws[f"J{r}"].fill = fill(YELLOW)
        for col in "IJK":
            ws[f"{col}{r}"].border = thin
    ws["J6"].fill = fill(GREEN_LIGHT)
    for col in "IJK":
        ws[f"{col}6"].border = thin

    merge_write(ws, "A16:G16", "1. Мониторинг предыдущих решений", font=font(True, 12, WHITE), fill=fill(BLUE), alignment=align("left", "center"))
    merge_write(ws, "A18:G18", "2. Решения текущей встречи · поставленные задачи", font=font(True, 12, WHITE), fill=fill(ORANGE), alignment=align("left", "center"))
    merge_write(ws, "A20:G20", "3. Обсуждаемые темы и выводы · блоки и тезисы", font=font(True, 12, WHITE), fill=fill(BLUE), alignment=align("left", "center"))


if __name__ == "__main__":
    main()
