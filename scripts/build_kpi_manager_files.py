#!/usr/bin/env python3
"""Generate KPI tracking Excel files for sales managers."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT_DIR = Path("/workspace/kpi_managers")
ARTIFACTS_DIR = Path("/opt/cursor/artifacts")

# Colors
BLUE_HEADER = "1F4E79"
BLUE_LIGHT = "D6E3F0"
YELLOW_INPUT = "FFF2CC"
GREEN_OK = "C6EFCE"
RED_BAD = "FFC7CE"
ORANGE_MID = "FCE4D6"
GRAY = "F2F2F2"
WHITE = "FFFFFF"
DARK = "1A1A1A"

thin = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


MANAGERS = [
    {
        "filename": "KPI_Кургузов_Данил.xlsx",
        "name": "Кургузов Данил",
        "role": "Менеджер по продажам",
        "kpis": [
            {
                "name": "Демо по Кабинету сотрудника",
                "unit": "шт",
                "plan": 10,
                "bonus": 5000,
                "kind": "number",
                "hint": "Количество проведённых демо",
            },
            {
                "name": "Подключения Кабинета сотрудника по пилоту",
                "unit": "шт",
                "plan": 3,
                "bonus": 5000,
                "kind": "number",
                "hint": "Количество подключений по пилоту",
            },
            {
                "name": "Успешные звонки в месяц",
                "unit": "шт",
                "plan": 350,
                "bonus": 5000,
                "kind": "number",
                "hint": "Кол-во успешных звонков",
            },
            {
                "name": "Заполнение показателей, ведение сделок, канбан и таблицы",
                "unit": "да/нет",
                "plan": 1,
                "bonus": 2500,
                "kind": "yesno",
                "hint": "1 = всё в порядке, 0 = нет",
            },
            {
                "name": "Соблюдение скриптов в звонках",
                "unit": "да/нет",
                "plan": 1,
                "bonus": 2500,
                "kind": "yesno",
                "hint": "1 = соблюдаются, 0 = нет",
            },
        ],
    },
    {
        "filename": "KPI_Оглоблина_Софья.xlsx",
        "name": "Оглоблина Софья",
        "role": "Менеджер по продажам",
        "kpis": [
            {
                "name": "Демо по Кабинету сотрудника",
                "unit": "шт",
                "plan": 10,
                "bonus": 5000,
                "kind": "number",
                "hint": "Количество проведённых демо",
            },
            {
                "name": "Подключения Доки.Логистика (демо-период) + контроль ранее подключённых",
                "unit": "шт",
                "plan": 5,
                "bonus": 5000,
                "kind": "number",
                "hint": "Новые подключения на демо + контроль старых клиентов",
            },
            {
                "name": "Успешные звонки в месяц",
                "unit": "шт",
                "plan": 350,
                "bonus": 5000,
                "kind": "number",
                "hint": "Кол-во успешных звонков",
            },
            {
                "name": "Заполнение показателей, ведение сделок, канбан и таблицы",
                "unit": "да/нет",
                "plan": 1,
                "bonus": 2500,
                "kind": "yesno",
                "hint": "1 = всё в порядке, 0 = нет",
            },
            {
                "name": "Соблюдение скриптов в звонках",
                "unit": "да/нет",
                "plan": 1,
                "bonus": 2500,
                "kind": "yesno",
                "hint": "1 = соблюдаются, 0 = нет",
            },
        ],
    },
    {
        "filename": "KPI_Юнусова_Юлиана.xlsx",
        "name": "Юнусова Юлиана",
        "role": "Менеджер по продажам / центр компетенций ЭПД",
        "kpis": [
            {
                "name": "Подключение сервиса 1С-ЭПД (пакеты титулов от 1000 шт)",
                "unit": "шт",
                "plan": 25,
                "bonus": 10000,
                "kind": "number",
                "hint": "Продажа пакетов титулов от 1000 шт",
            },
            {
                "name": "Консультация менеджеров (центр компетенций ЭПД)",
                "unit": "да/нет",
                "plan": 1,
                "bonus": 10000,
                "kind": "yesno",
                "hint": "1 = консультации выполнены, 0 = нет",
            },
        ],
    },
]


def style_cell(cell, *, fill=None, font=None, align=None, border=True, number_format=None):
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = thin
    if number_format:
        cell.number_format = number_format


def build_workbook(manager: dict) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "KPI месяц"

    # Column widths
    widths = {
        "A": 4,
        "B": 62,
        "C": 10,
        "D": 12,
        "E": 14,
        "F": 14,
        "G": 14,
        "H": 16,
        "I": 18,
        "J": 28,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 10
    ws.row_dimensions[5].height = 24

    # Title
    ws.merge_cells("B1:I1")
    ws["B1"] = f"КИПИАЙ — {manager['name']}"
    style_cell(
        ws["B1"],
        fill=BLUE_HEADER,
        font=Font(name="Calibri", size=18, bold=True, color=WHITE),
        align=Alignment(horizontal="left", vertical="center"),
        border=False,
    )
    for col in range(3, 10):
        style_cell(
            ws.cell(1, col),
            fill=BLUE_HEADER,
            border=False,
        )

    ws.merge_cells("B2:I2")
    ws["B2"] = manager["role"]
    style_cell(
        ws["B2"],
        fill=BLUE_LIGHT,
        font=Font(name="Calibri", size=11, italic=True, color=DARK),
        align=Alignment(vertical="center"),
        border=False,
    )
    for col in range(3, 10):
        style_cell(ws.cell(2, col), fill=BLUE_LIGHT, border=False)

    # Period / month input
    ws["B3"] = "Период (месяц):"
    style_cell(
        ws["B3"],
        font=Font(name="Calibri", size=11, bold=True),
        align=Alignment(vertical="center", horizontal="right"),
        border=False,
    )
    ws.merge_cells("C3:D3")
    ws["C3"] = ""
    style_cell(
        ws["C3"],
        fill=YELLOW_INPUT,
        font=Font(name="Calibri", size=11, bold=True),
        align=Alignment(horizontal="center", vertical="center"),
    )
    style_cell(ws["D3"], fill=YELLOW_INPUT)
    ws["E3"] = "← впишите месяц, например: Сентябрь 2026"
    style_cell(
        ws["E3"],
        font=Font(name="Calibri", size=9, italic=True, color="666666"),
        border=False,
    )
    ws.merge_cells("E3:I3")

    # Headers row 5
    headers = [
        ("B5", "Показатель КИПИАЙ"),
        ("C5", "Ед."),
        ("D5", "План"),
        ("E5", "Факт"),
        ("F5", "% выполнения"),
        ("G5", "Бонус, ₽"),
        ("H5", "Начислено, ₽"),
        ("I5", "Статус"),
        ("J5", "Комментарий менеджера"),
    ]
    for cell_ref, title in headers:
        cell = ws[cell_ref]
        cell.value = title
        style_cell(
            cell,
            fill=BLUE_HEADER,
            font=Font(name="Calibri", size=11, bold=True, color=WHITE),
            align=Alignment(horizontal="center", vertical="center", wrap_text=True),
        )

    first_data_row = 6
    yesno_rows = []

    for i, kpi in enumerate(manager["kpis"]):
        row = first_data_row + i
        ws.row_dimensions[row].height = 36

        # Name
        ws.cell(row, 2, kpi["name"])
        style_cell(
            ws.cell(row, 2),
            fill=WHITE if i % 2 == 0 else GRAY,
            font=Font(name="Calibri", size=11),
            align=Alignment(vertical="center", wrap_text=True),
        )

        # Unit
        ws.cell(row, 3, kpi["unit"])
        style_cell(
            ws.cell(row, 3),
            fill=WHITE if i % 2 == 0 else GRAY,
            font=Font(name="Calibri", size=10),
            align=Alignment(horizontal="center", vertical="center"),
        )

        # Plan
        ws.cell(row, 4, kpi["plan"])
        style_cell(
            ws.cell(row, 4),
            fill=WHITE if i % 2 == 0 else GRAY,
            font=Font(name="Calibri", size=11, bold=True),
            align=Alignment(horizontal="center", vertical="center"),
            number_format="0",
        )

        # Fact — INPUT
        fact_cell = ws.cell(row, 5, None)
        style_cell(
            fact_cell,
            fill=YELLOW_INPUT,
            font=Font(name="Calibri", size=12, bold=True, color="9C5700"),
            align=Alignment(horizontal="center", vertical="center"),
            number_format="0",
        )
        fact_cell.comment = None
        from openpyxl.comments import Comment

        fact_cell.comment = Comment(kpi["hint"], "KPI", width=240, height=60)

        # % completion = Fact / Plan, capped display via formula
        pct_cell = ws.cell(row, 6)
        pct_cell.value = f'=IF(OR(E{row}="",D{row}=0),"",MIN(1,E{row}/D{row}))'
        style_cell(
            pct_cell,
            fill=WHITE if i % 2 == 0 else GRAY,
            font=Font(name="Calibri", size=11, bold=True),
            align=Alignment(horizontal="center", vertical="center"),
            number_format="0%",
        )

        # Max bonus
        ws.cell(row, 7, kpi["bonus"])
        style_cell(
            ws.cell(row, 7),
            fill=WHITE if i % 2 == 0 else GRAY,
            font=Font(name="Calibri", size=11),
            align=Alignment(horizontal="center", vertical="center"),
            number_format="#,##0",
        )

        # Earned bonus: proportional, capped at 100% of max bonus
        # For yes/no: only full if fact >= 1
        earned = ws.cell(row, 8)
        if kpi["kind"] == "yesno":
            earned.value = f'=IF(E{row}="","",IF(E{row}>=1,G{row},0))'
        else:
            earned.value = f'=IF(E{row}="","",ROUND(MIN(1,E{row}/D{row})*G{row},0))'
        style_cell(
            earned,
            fill=WHITE if i % 2 == 0 else GRAY,
            font=Font(name="Calibri", size=11, bold=True),
            align=Alignment(horizontal="center", vertical="center"),
            number_format="#,##0",
        )

        # Status
        status = ws.cell(row, 9)
        status.value = (
            f'=IF(E{row}="","ожидает ввода",'
            f'IF(E{row}>=D{row},"выполнено","не выполнено"))'
        )
        style_cell(
            status,
            fill=WHITE if i % 2 == 0 else GRAY,
            font=Font(name="Calibri", size=10),
            align=Alignment(horizontal="center", vertical="center", wrap_text=True),
        )

        # Comment input
        comment_cell = ws.cell(row, 10, "")
        style_cell(
            comment_cell,
            fill=YELLOW_INPUT,
            font=Font(name="Calibri", size=10),
            align=Alignment(vertical="center", wrap_text=True),
        )

        if kpi["kind"] == "yesno":
            yesno_rows.append(row)

    last_data_row = first_data_row + len(manager["kpis"]) - 1

    # Data validation for yes/no
    if yesno_rows:
        dv = DataValidation(
            type="list",
            formula1='"0,1"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Ввод",
            error="Введите 0 (нет) или 1 (да)",
            promptTitle="Да / Нет",
            prompt="1 = выполнено, 0 = не выполнено",
            showInputMessage=True,
        )
        ws.add_data_validation(dv)
        for r in yesno_rows:
            dv.add(ws.cell(r, 5))

    # Conditional formatting for status / %
    green_fill = PatternFill("solid", fgColor=GREEN_OK)
    red_fill = PatternFill("solid", fgColor=RED_BAD)
    orange_fill = PatternFill("solid", fgColor=ORANGE_MID)
    ws.conditional_formatting.add(
        f"F{first_data_row}:F{last_data_row}",
        CellIsRule(operator="greaterThanOrEqual", formula=["1"], fill=green_fill),
    )
    ws.conditional_formatting.add(
        f"F{first_data_row}:F{last_data_row}",
        CellIsRule(operator="between", formula=["0.7", "0.999"], fill=orange_fill),
    )
    ws.conditional_formatting.add(
        f"F{first_data_row}:F{last_data_row}",
        CellIsRule(operator="lessThan", formula=["0.7"], fill=red_fill),
    )
    ws.conditional_formatting.add(
        f"I{first_data_row}:I{last_data_row}",
        FormulaRule(
            formula=[f'I{first_data_row}="выполнено"'],
            fill=green_fill,
        ),
    )
    ws.conditional_formatting.add(
        f"I{first_data_row}:I{last_data_row}",
        FormulaRule(
            formula=[f'I{first_data_row}="не выполнено"'],
            fill=red_fill,
        ),
    )

    # Summary block
    summary_row = last_data_row + 2
    ws.row_dimensions[summary_row].height = 26
    ws.merge_cells(f"B{summary_row}:D{summary_row}")
    ws.cell(summary_row, 2, "ИТОГО за месяц")
    style_cell(
        ws.cell(summary_row, 2),
        fill=BLUE_HEADER,
        font=Font(name="Calibri", size=12, bold=True, color=WHITE),
        align=Alignment(horizontal="left", vertical="center"),
    )
    for col in (3, 4):
        style_cell(ws.cell(summary_row, col), fill=BLUE_HEADER, border=True)

    ws.cell(summary_row, 5, "Макс. бонус:")
    style_cell(
        ws.cell(summary_row, 5),
        fill=BLUE_LIGHT,
        font=Font(name="Calibri", size=10, bold=True),
        align=Alignment(horizontal="right", vertical="center"),
    )
    ws.cell(summary_row, 6, f"=SUM(G{first_data_row}:G{last_data_row})")
    style_cell(
        ws.cell(summary_row, 6),
        fill=BLUE_LIGHT,
        font=Font(name="Calibri", size=12, bold=True),
        align=Alignment(horizontal="center", vertical="center"),
        number_format="#,##0 ₽",
    )

    ws.cell(summary_row, 7, "Начислено:")
    style_cell(
        ws.cell(summary_row, 7),
        fill=GREEN_OK,
        font=Font(name="Calibri", size=10, bold=True),
        align=Alignment(horizontal="right", vertical="center"),
    )
    ws.cell(summary_row, 8, f"=SUM(H{first_data_row}:H{last_data_row})")
    style_cell(
        ws.cell(summary_row, 8),
        fill=GREEN_OK,
        font=Font(name="Calibri", size=14, bold=True),
        align=Alignment(horizontal="center", vertical="center"),
        number_format="#,##0 ₽",
    )

    # Count of KPIs where fact >= plan (pairwise via status column)
    ws.cell(
        summary_row,
        9,
        f'=IF(COUNTBLANK(E{first_data_row}:E{last_data_row})=ROWS(E{first_data_row}:E{last_data_row}),'
        f'"заполните факт",'
        f'IF(COUNTIF(I{first_data_row}:I{last_data_row},"выполнено")=ROWS(E{first_data_row}:E{last_data_row}),'
        f'"все KPI выполнены","есть недовыполнение"))',
    )
    style_cell(
        ws.cell(summary_row, 9),
        fill=BLUE_LIGHT,
        font=Font(name="Calibri", size=10, bold=True),
        align=Alignment(horizontal="center", vertical="center", wrap_text=True),
    )
    style_cell(ws.cell(summary_row, 10), fill=BLUE_LIGHT)

    # Overall % of plan (average of individual %)
    pct_row = summary_row + 1
    ws.merge_cells(f"B{pct_row}:D{pct_row}")
    ws.cell(pct_row, 2, "Средний % выполнения плана")
    style_cell(
        ws.cell(pct_row, 2),
        fill=GRAY,
        font=Font(name="Calibri", size=11, bold=True),
        align=Alignment(vertical="center"),
    )
    for col in (3, 4):
        style_cell(ws.cell(pct_row, col), fill=GRAY)
    ws.merge_cells(f"E{pct_row}:F{pct_row}")
    ws.cell(
        pct_row,
        5,
        f'=IF(COUNT(E{first_data_row}:E{last_data_row})=0,"",AVERAGE(F{first_data_row}:F{last_data_row}))',
    )
    style_cell(
        ws.cell(pct_row, 5),
        fill=GRAY,
        font=Font(name="Calibri", size=14, bold=True),
        align=Alignment(horizontal="center", vertical="center"),
        number_format="0.0%",
    )
    style_cell(ws.cell(pct_row, 6), fill=GRAY)
    for col in range(7, 11):
        style_cell(ws.cell(pct_row, col), fill=GRAY, border=True)

    # Instructions sheet
    info = wb.create_sheet("Инструкция")
    info.column_dimensions["A"].width = 100
    info["A1"] = "Как пользоваться файлом"
    info["A1"].font = Font(name="Calibri", size=16, bold=True, color=BLUE_HEADER)
    lines = [
        "",
        "1. На листе «KPI месяц» в жёлтой ячейке укажите период (например: Сентябрь 2026).",
        "2. В колонке «Факт» (жёлтые ячейки) впишите фактические показатели за месяц.",
        "3. Для показателей «да/нет» введите 1 (выполнено) или 0 (не выполнено) — есть выпадающий список.",
        "4. Колонки «% выполнения», «Начислено, ₽» и «Статус» считаются автоматически.",
        "5. Бонус по числовым KPI начисляется пропорционально выполнению плана (не более 100%).",
        "6. По KPI «да/нет» бонус начисляется полностью только при факте = 1.",
        "7. В колонке «Комментарий менеджера» можно кратко пояснить результат.",
        "8. Итоговая сумма бонуса — в зелёной ячейке «Начислено» внизу таблицы.",
        "",
        "Не изменяйте формулы в белых колонках — только жёлтые ячейки для ввода.",
    ]
    for idx, line in enumerate(lines, start=2):
        info.cell(idx, 1, line)
        info.cell(idx, 1).font = Font(name="Calibri", size=11)

    # Freeze panes
    ws.freeze_panes = "B6"
    ws.print_title_rows = "1:5"

    return wb


def main():
    import shutil

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    ARTIFACTS_DIR.mkdir(parents=True, exist_ok=True)

    created = []
    for manager in MANAGERS:
        wb = build_workbook(manager)
        path = OUT_DIR / manager["filename"]
        wb.save(path)
        art = ARTIFACTS_DIR / manager["filename"]
        shutil.copy2(path, art)
        created.append(path)
        print(f"Created: {path}")

    print(f"Done: {len(created)} files")


if __name__ == "__main__":
    main()
