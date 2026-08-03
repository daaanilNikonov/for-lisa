#!/usr/bin/env python3
"""Build sales funnel flowchart Excel for managers."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "Блок-схема работы с клиентом.xlsx"

DARK = "1A1A1A"
WHITE = "FFFFFF"
BLUE = "26A6E0"
HEADER = "1E3A4C"
SOFT = "F2F7FA"
ACCENT = "D6EEF8"
OK = "EAF7EE"
WARN = "FFF4E5"
PINK = "FDECEC"
PURPLE = "F3EEF8"

# Stage colors
C1 = "26A6E0"
C2 = "2E8B57"
C3 = "D97706"
C4 = "B45309"

thin = Side(style="thin", color="C5D5E0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
FONT = "Calibri"


def fnt(size=11, bold=False, color=DARK):
    return Font(name=FONT, size=size, bold=bold, color=color)


def fill(c):
    return PatternFill("solid", fgColor=c)


def al(h="left", v="top", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def style_cell(cell, bg=None, font=None, alignment=None):
    cell.border = BORDER
    if bg:
        cell.fill = fill(bg)
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment


def build_flowchart(wb: Workbook):
    ws = wb.active
    ws.title = "Блок-схема"

    ws.merge_cells("A1:D1")
    ws["A1"] = "БЛОК-СХЕМА РАБОТЫ С КЛИЕНТОМ · «1С:Кабинет сотрудника»"
    for c in range(1, 5):
        style_cell(ws.cell(1, c), HEADER, fnt(16, True, WHITE), al("center", "center"))
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:D2")
    ws["A2"] = (
        "Путь клиента: знакомство → презентация → демо со специалистом → счёт на 10 кабинетов (пилот). "
        "В каждой колонке: цель, что узнать, что рассказать, следующий шаг."
    )
    for c in range(1, 5):
        style_cell(ws.cell(2, c), BLUE, fnt(10, False, WHITE), al("center", "center"))
    ws.row_dimensions[2].height = 36

    headers = [
        ("① ПЕРВОЕ ЗНАКОМСТВО", C1),
        ("② ПРЕДДЕМОНСТРАЦИЯ", C2),
        ("③ ДЕМОНСТРАЦИЯ", C3),
        ("④ ЗАКРЫТИЕ НА СЧЁТ", C4),
    ]
    for i, (title, color) in enumerate(headers, 1):
        cell = ws.cell(3, i, title)
        style_cell(cell, color, fnt(12, True, WHITE), al("center", "center"))
    ws.row_dimensions[3].height = 28

    # Flow line
    ws.merge_cells("A4:D4")
    ws["A4"] = "➔  звонок по скрипту  ➔  презентация и запись на демо  ➔  демо со специалистом  ➔  счёт на 10 кабинетов  ➔"
    for c in range(1, 5):
        style_cell(ws.cell(4, c), SOFT, fnt(10, True, DARK), al("center", "center"))
    ws.row_dimensions[4].height = 22

    # Format / channel
    formats = [
        "Формат: телефонный звонок\nМатериал: скрипт КабС / ДОКИ / Смартвей",
        "Формат: созвон / встреча\nМатериал: презентация «Продвижение 1С:Кабинет сотрудника»",
        "Формат: онлайн-демо\nМатериал: регламент проведения демонстраций",
        "Формат: звонок после демо\nМатериал: коммерция, счёт на 10 кабинетов",
    ]
    for i, text in enumerate(formats, 1):
        style_cell(ws.cell(5, i, text), ACCENT, fnt(9, False, DARK), al("center", "center"))
    ws.row_dimensions[5].height = 42

    # Goal
    for i, text in enumerate(
        [
            "ЦЕЛЬ\nПолучить 2–5 минут, понять боль, заинтересовать и договориться о следующем шаге",
            "ЦЕЛЬ\nПоказать ценность под роль клиента, снять страхи и записать на демо со специалистом",
            "ЦЕЛЬ\nПоказать сервис «вживую» на задачах клиента и получить «да, это нам подходит»",
            "ЦЕЛЬ\nПеревести интерес в пилот: счёт на 10 кабинетов без давления и спешки",
        ],
        1,
    ):
        style_cell(ws.cell(6, i, text), WARN, fnt(10, True, DARK), al("left", "top"))
    ws.row_dimensions[6].height = 68

    # Learn
    learn_label_bg = "E8F4FB"
    learns = [
        "УЗНАТЬ У КЛИЕНТА\n"
        "• Кто на линии: руководитель / бухгалтер / кадровик / IT\n"
        "• Как сейчас: бумага или уже электронно\n"
        "• Как выдают расчётные листки, есть ли подтверждение ознакомления\n"
        "• Как подают заявления, отпуска, справки; есть ли удалёнка\n"
        "• Сколько сотрудников примерно в компании\n"
        "• Для бухгалтера: как обмениваются документами с контрагентами (заход в ДОКИ)\n"
        "• Бывают ли командировки и как часто (заход в Смартвей)",
        "УЗНАТЬ У КЛИЕНТА\n"
        "• Кто придёт на демо (лучше бухгалтер + кадровик, при необходимости IT)\n"
        "• Что важнее всего увидеть: расчётки / отпуска / заявления / удалёнка / безопасность\n"
        "• Сколько сотрудников готовы взять на пилот (цель — 10 кабинетов)\n"
        "• Что смущает: дорого, некогда, сотрудники не разберутся, IT загружен\n"
        "• Удобные дата и время демо",
        "УЗНАТЬ У КЛИЕНТА\n"
        "• Какой сценарий показать первым (под боль с прошлого этапа)\n"
        "• Нужны ли IT-вопросы: хранение в базе, ФСТЭК, локальная версия\n"
        "• После показа: «Это решает вашу задачу?»\n"
        "• Что мешает стартовать?\n"
        "• Кто принимает решение по счёту",
        "УЗНАТЬ У КЛИЕНТА\n"
        "• Что понравилось на демо, что осталось под вопросом\n"
        "• Какой отдел / каких 10 сотрудников берём в пилот\n"
        "• С чего стартуем: расчётные листки или сразу отпуска и заявления\n"
        "• Кто подписывает / оплачивает счёт\n"
        "• Нужен ли IT на подключении, удобная дата старта",
    ]
    for i, text in enumerate(learns, 1):
        style_cell(ws.cell(7, i, text), learn_label_bg, fnt(9, False, DARK), al("left", "top"))
    ws.row_dimensions[7].height = 155

    # Tell
    tells = [
        "РАССКАЗАТЬ КЛИЕНТУ\n"
        "• Мы из «Форуса», ваш партнёр по 1С\n"
        "• ДОКИ: по ИТС ПРОФ — 3 месяца безлимита бесплатно, можно не продлевать\n"
        "• Кабинет сотрудника: как Госуслуги для сотрудников\n"
        "• Можно начать с одного отдела, меньше 30 ₽/мес на человека\n"
        "• Ст. 136 и 173 ТК — лучше подготовиться спокойно\n"
        "• Смартвей — если есть командировки (передать коллеге)\n"
        "• Предложить 15-минутный показ / следующий контакт",
        "РАССКАЗАТЬ КЛИЕНТУ\n"
        "• Зачем сервис именно им — по их боли\n"
        "• Выгоды: меньше рутины, расчётки по закону, данные сразу в 1С\n"
        "• Документы в их базе, настройку берём на себя\n"
        "• Мягкий вход: один отдел / 10 сотрудников\n"
        "• Ориентир цены: ~3 360 ₽/год за 10 кабинетов\n"
        "• Как пройдёт демо: дата, время, кто будет со стороны Форус\n"
        "• При необходимости — 45 дней тест-драйва (если идёт акция)",
        "РАССКАЗАТЬ / ПОКАЗАТЬ\n"
        "• Менеджер: знакомит со специалистом, коротко говорит план демо\n"
        "• Специалист показывает на примере:\n"
        "  — личный кабинет сотрудника\n"
        "  — расчётный листок с ознакомлением\n"
        "  — заявление → сразу в 1С / приказ\n"
        "  — согласование руководителем\n"
        "• Спокойно отвечаем на вопросы по безопасности и подписи\n"
        "• Не давим на покупку на самой демонстрации",
        "РАССКАЗАТЬ КЛИЕНТУ\n"
        "• Предлагаем пилот на 10 кабинетов — без перевода всей компании сразу\n"
        "• Стоимость пилота понятная и небольшая (~3 360 ₽/год)\n"
        "• Настройку и обучение берём на себя\n"
        "• Есть шаблоны документов для перехода на КЭДО\n"
        "• После пилота легко расширить на всю компанию\n"
        "• Когда КЭДО станет обязательным — вы уже будете готовы",
    ]
    for i, text in enumerate(tells, 1):
        style_cell(ws.cell(8, i, text), OK, fnt(9, False, DARK), al("left", "top"))
    ws.row_dimensions[8].height = 175

    # CTA / result
    ctas = [
        "СЛЕДУЮЩИЙ ШАГ\n"
        "Договориться о презентации / демо.\n"
        "Зафиксировать: ФИО, роль, телефон, e-mail, число сотрудников, дату следующего контакта.",
        "СЛЕДУЮЩИЙ ШАГ\n"
        "Записать на демо.\n"
        "Письмо клиенту: дата, время, ссылка.\n"
        "Ту же ссылку — специалисту.\n"
        "За 1 день и за 1 час — подтвердить участие.",
        "СЛЕДУЮЩИЙ ШАГ\n"
        "Менеджер после вводной может выйти и позже собрать ОС.\n"
        "В течение 3 дней — созвон: впечатления + предложение пилота на 10 кабинетов.",
        "СЛЕДУЮЩИЙ ШАГ\n"
        "Выставить счёт на 10 кабинетов.\n"
        "Согласовать дату подключения.\n"
        "Передать на внедрение.\n"
        "Через 2–4 недели — «как пилот, расширяем?»",
    ]
    for i, text in enumerate(ctas, 1):
        style_cell(ws.cell(9, i, text), PINK if i == 4 else ACCENT, fnt(9, True, DARK), al("left", "top"))
    ws.row_dimensions[9].height = 95

    # Checklist row labels as separate mini sheet content - tips
    ws.merge_cells("A10:D10")
    ws["A10"] = "КОРОТКАЯ ПАМЯТКА"
    for c in range(1, 5):
        style_cell(ws.cell(10, c), HEADER, fnt(11, True, WHITE), al("left", "center"))
    ws.row_dimensions[10].height = 22

    tips = [
        "Не продавать всё сразу. Сначала боль, потом сервис. Кадровику ДОКИ не продавливать.",
        "Перед демо сверить ссылку у клиента и у специалиста. Без подтверждения — высокий риск срыва.",
        "Демо — про пользу, не про прайс. Цену и счёт — на следующем контакте после ОС.",
        "Пилот = низкий риск для клиента. 10 кабинетов — лёгкий вход, потом вся компания.",
    ]
    for i, text in enumerate(tips, 1):
        style_cell(ws.cell(11, i, text), "FFF8E7", fnt(9, False, DARK), al("left", "top"))
    ws.row_dimensions[11].height = 55

    for i in range(1, 5):
        ws.column_dimensions[get_column_letter(i)].width = 36
    ws.freeze_panes = "A5"
    ws.print_title_rows = "1:3"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1


def build_timeline(wb: Workbook):
    """Vertical process view — easier to follow step by step."""
    ws = wb.create_sheet("По шагам")

    ws.merge_cells("A1:D1")
    ws["A1"] = "СКРИПТ-БЛОК СХЕМА · ПО ШАГАМ"
    for c in range(1, 5):
        style_cell(ws.cell(1, c), HEADER, fnt(16, True, WHITE), al("center", "center"))
    ws.row_dimensions[1].height = 28

    headers = ["Этап", "Что узнать", "Что рассказать / сделать", "Результат этапа"]
    for i, h in enumerate(headers, 1):
        style_cell(ws.cell(2, i, h), BLUE, fnt(11, True, WHITE), al("center", "center"))
    ws.row_dimensions[2].height = 22

    rows = [
        (
            "1. Первое знакомство\n(скрипт звонка)",
            "Кто ЛПР; как устроен кадровый документооборот; расчётные листки; отпуска; удалёнка; число сотрудников; ЭДО с контрагентами; командировки",
            "Представиться от «Форуса». ДОКИ через ИТС ПРОФ. Кабинет сотрудника — как Госуслуги, старт с одного отдела, меньше 30 ₽/чел. При необходимости — Смартвей. Предложить показ.",
            "Договорённость о презентации / демо.\nКонтакты и роль зафиксированы.",
            C1,
        ),
        (
            "2. Преддемонстрация\n(презентация)",
            "Кто будет на демо; какой сценарий важнее; сколько кабинетов на пилот; возражения; удобные дата и время",
            "Пройтись по презентации под боль клиента. Показать выгоды и мягкий вход на 10 человек. Объяснить, как пройдёт демо. Отправить письмо: дата, время, ссылка. Ссылку — специалисту. Подтвердить за 1 день и за 1 час.",
            "Клиент записан и подтвердил участие.\nСпециалист в курсе и со ссылкой.",
            C2,
        ),
        (
            "3. Демонстрация\n(со специалистом)",
            "Что показать первым; вопросы по безопасности/IT; после демо — решает ли задачу; что мешает старту; кто согласует оплату",
            "За 5 минут менеджер проверяет ссылку, знакомит клиента со специалистом, озвучивает план. Специалист показывает кабинет, расчётки, заявления в 1С, согласование. Менеджер может выйти и позже собрать ОС (до 3 дней).",
            "Клиент увидел сервис на своих задачах.\nЕсть обратная связь и готовность обсудить пилот.",
            C3,
        ),
        (
            "4. Закрытие на счёт\n(пилот 10 кабинетов)",
            "Что понравилось; состав пилота на 10 человек; с чего стартуем; кто платит; дата подключения",
            "Предложить пилот на 10 кабинетов (~3 360 ₽/год). Настройку и обучение берём на себя. Выставить счёт. Согласовать старт. Через 2–4 недели — разговор о расширении на компанию.",
            "Счёт выставлен / оплачен.\nПилот запущен. Есть план follow-up.",
            C4,
        ),
    ]

    for r, (stage, learn, tell, result, color) in enumerate(rows, 3):
        vals = [stage, learn, tell, result]
        bgs = [color, SOFT, OK, WARN]
        fonts = [fnt(10, True, WHITE), fnt(9), fnt(9), fnt(9, True)]
        for c, (val, bg, font) in enumerate(zip(vals, bgs, fonts), 1):
            style_cell(ws.cell(r, c, val), bg, font, al("left", "top"))
        ws.row_dimensions[r].height = 120

    # Arrow helper
    ws.merge_cells("A7:D7")
    ws["A7"] = (
        "Логика воронки: не перескакивать этапы. "
        "Сначала боль и интерес → потом презентация → потом «потрогать» сервис → потом маленький счёт на пилот."
    )
    for c in range(1, 5):
        style_cell(ws.cell(7, c), "FFF8E7", fnt(10, True, DARK), al("left", "center"))
    ws.row_dimensions[7].height = 36

    # Mini checklist after demo close
    ws.merge_cells("A8:D8")
    ws["A8"] = "ЧЕК-ЛИСТ ПЕРЕД СЧЁТОМ НА 10 КАБИНЕТОВ"
    for c in range(1, 5):
        style_cell(ws.cell(8, c), HEADER, fnt(11, True, WHITE), al("left", "center"))

    checks = [
        "☐ Демо проведено, ОС собрана",
        "☐ Согласован пилот: 10 сотрудников / отдел",
        "☐ Понятен старт: расчётки или отпуска",
        "☐ Есть ЛПР на оплату и дата подключения",
    ]
    for i, t in enumerate(checks, 1):
        style_cell(ws.cell(9, i, t), ACCENT, fnt(9, True, DARK), al("left", "center"))
    ws.row_dimensions[9].height = 30

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 32
    ws.freeze_panes = "A3"


def main():
    wb = Workbook()
    build_flowchart(wb)
    build_timeline(wb)
    wb.save(OUT)
    print(f"Saved: {OUT}")


if __name__ == "__main__":
    main()
