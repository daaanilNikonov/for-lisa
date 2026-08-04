#!/usr/bin/env python3
"""
Сборка инвестиционной модели «1С:Кабинет сотрудника».

Берёт исходный Excel, обогащает базу (ССЧР ФНС + ОКВЭД ЕГРЮЛ),
пересобирает листы с единой точкой ввода и связанными формулами.
"""

from __future__ import annotations

import json
import re
import shutil
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "Финансовый_план_Кабинет_сотрудника_ИСПРАВЛЕННЫЙ(АвтоматическиВосстановлено).xlsx"
OUT = ROOT / "Финансовый_план_Кабинет_сотрудника_ИНВЕСТИЦИОННАЯ_МОДЕЛЬ.xlsx"
DATA = ROOT / "data"

GREEN = PatternFill("solid", fgColor="C6EFCE")
GRAY = PatternFill("solid", fgColor="D9D9D9")
RED = PatternFill("solid", fgColor="FFC7CE")
BLUE_HDR = PatternFill("solid", fgColor="1F4E79")
LIGHT_BLUE = PatternFill("solid", fgColor="D6EAF8")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
ORANGE = PatternFill("solid", fgColor="FCE4D6")
WHITE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="1F4E79")
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color="1F4E79")
BOLD = Font(name="Calibri", bold=True, size=11)
NORMAL = Font(name="Calibri", size=10)
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

# Целевые префиксы ОКВЭД (ICP для КЭДО / Кабинет сотрудника)
TARGET_OKVED_PREFIXES = [
    ("45", "G — торговля авто"),
    ("46", "G — оптовая торговля"),
    ("47", "G — розничная торговля"),
    ("10", "C — пищевое производство"),
    ("11", "C — напитки"),
    ("12", "C — табак"),
    ("13", "C — текстиль"),
    ("14", "C — одежда"),
    ("15", "C — кожа"),
    ("16", "C — деревообработка"),
    ("17", "C — бумага"),
    ("18", "C — полиграфия"),
    ("19", "C — кокс/нефтепродукты"),
    ("20", "C — химия"),
    ("21", "C — лекарства"),
    ("22", "C — резина/пластик"),
    ("23", "C — неметаллы"),
    ("24", "C — металлургия"),
    ("25", "C — металлоизделия"),
    ("26", "C — компьютеры/электроника"),
    ("27", "C — электрооборудование"),
    ("28", "C — машины"),
    ("29", "C — автотранспортные средства"),
    ("30", "C — прочий транспорт"),
    ("31", "C — мебель"),
    ("32", "C — прочие готовые изделия"),
    ("33", "C — ремонт/монтаж машин"),
    ("41", "F — строительство зданий"),
    ("42", "F — инженерия"),
    ("43", "F — строительные работы"),
    ("49", "H — сухопутный транспорт"),
    ("52", "H — складирование/логистика"),
    ("55", "I — гостиницы"),
    ("56", "I — общепит"),
    ("68", "L — недвижимость"),
    ("78", "N — подбор персонала"),
    ("80", "N — охрана"),
    ("81", "N — обслуживание зданий"),
    ("82", "N — административные услуги"),
    ("64", "K — финансы"),
    ("65", "K — страхование"),
    ("66", "K — вспомогательные финансы"),
    ("61", "J — телеком"),
    ("62", "J — IT"),
    ("63", "J — информационные услуги"),
    ("85", "P — образование"),
    ("86", "Q — здравоохранение"),
]

TARIFFS = [
    # пакет, сотрудников, розница/мес, розница/год, закупка/год
    ("До 10", 10, 280, 3360, 1680),
    ("До 25", 25, 700, 8400, 4200),
    ("До 50", 50, 1400, 16800, 8400),
    ("До 100", 100, 2800, 33600, 16800),
    ("До 200", 200, 5200, 62400, 31200),
    ("До 500", 500, 12000, 144000, 72000),
    ("До 2000", 2000, 48000, 528000, 264000),
    ("До 5000", 5000, 115000, 1260000, 630000),
]


def style_header_row(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = BLUE_HDR
        cell.font = WHITE_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN


def style_input(cell):
    cell.fill = GREEN
    cell.font = NORMAL
    cell.border = THIN


def style_formula(cell):
    cell.fill = GRAY
    cell.font = NORMAL
    cell.border = THIN


def style_control(cell):
    cell.fill = RED
    cell.font = BOLD
    cell.border = THIN


def set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def is_target_okved(code: str | None, all_codes: str | None) -> str:
    text = " ".join(x for x in [code or "", all_codes or ""] if x)
    if not text.strip():
        return "Не определено"
    # normalize separators
    parts = re.split(r"[;\s,]+", text)
    for p in parts:
        p = p.strip()
        if not p:
            continue
        prefix2 = p[:2]
        for pref, _ in TARGET_OKVED_PREFIXES:
            if p.startswith(pref) or prefix2 == pref:
                return "Да"
    return "Нет"


def package_for_headcount(n):
    if n is None or n == "":
        return "", None
    try:
        n = float(n)
    except Exception:
        return "", None
    for name, max_emp, _m, year, _c in TARIFFS:
        if n <= max_emp:
            return name, year
    return "Индивидуально", None


def load_enrichment():
    sschr_path = DATA / "sschr_fns_2025.json"
    if not sschr_path.exists():
        sschr_path = Path("/tmp/sschr_index.json")
    sschr = json.loads(sschr_path.read_text(encoding="utf-8")) if sschr_path.exists() else {}

    okved = {}
    for p in [Path("/tmp/okved_index.json"), DATA / "okved_egrul.json"]:
        if p.exists():
            okved.update(json.loads(p.read_text(encoding="utf-8")))
    cache_dir = Path("/tmp/okved_cache")
    if cache_dir.exists():
        for f in cache_dir.glob("*.json"):
            try:
                item = json.loads(f.read_text(encoding="utf-8"))
                if item.get("inn"):
                    okved[item["inn"]] = item
            except Exception:
                pass
    return sschr, okved


def enrich_clients(wb, sschr: dict, okved: dict) -> dict:
    ws = wb["Данные клиентов"]
    # Ensure headers for new columns
    headers = {
        1: "Регномер",
        2: "Клиент",
        3: "ИНН / ЕДРПОУ из базы",
        4: "ИНН",
        5: "Тип ИНН",
        6: "Комментарий из исходной базы",
        7: "Численность сотрудников (ССЧР)",
        8: "Год ССЧР",
        9: "Основной ОКВЭД",
        10: "Все ОКВЭД",
        11: "Подходит по численности ≥22?",
        12: "Подходит по целевому ОКВЭД?",
        13: "Пакет апселла (по ССЧР)",
        14: "Потенциал пилота КП, ₽/год (10 каб. = 3360)",
        15: "Категория по численности",
        16: "Статус проверки",
        17: "Источник ССЧР",
        18: "Источник ОКВЭД",
        19: "Уникальный по ИНН",
        20: "Приоритет",
        21: "Сегмент",
        22: "ОКВЭД подходит?",
        23: "Численность подходит?",
        24: "ОКВЭД + численность (приоритет №1)",
    }
    for c, h in headers.items():
        ws.cell(4, c).value = h
    style_header_row(ws, 4, 24)

    ws["A1"] = "База клиентов — сегментация по ОКВЭД и численности"
    ws["A2"] = (
        "Автоматически: целевой ОКВЭД / 22+ / пересечение (приоритет №1). "
        "ССЧР — открытые данные ФНС; ОКВЭД — выписки ЕГРЮЛ. Зелёные поля можно уточнять вручную."
    )
    ws["A3"] = "Не удаляйте формулы в столбцах K–O и T–X. Ввод: G, H, I, J и комментарии."

    # First pass: uniqueness (reset column S — old file may contain stale flags)
    inn_rows: dict[str, list[int]] = {}
    max_row = ws.max_row
    for r in range(5, max_row + 1):
        ws.cell(r, 19).value = 0
        inn = str(ws.cell(r, 4).value or "").strip()
        if not inn:
            raw = ws.cell(r, 3).value
            if raw is not None:
                inn = re.sub(r"\D", "", str(raw))
                if len(inn) in (10, 12):
                    ws.cell(r, 4).value = inn
                else:
                    inn = ""
        if inn:
            inn_rows.setdefault(inn, []).append(r)

    stats = {
        "rows": 0,
        "with_inn": 0,
        "unique_inn": len(inn_rows),
        "sschr": 0,
        "okved": 0,
        "target_okved": 0,
        "ge22": 0,
        "hot": 0,
        "okved_only": 0,
        "ge22_only": 0,
        "ge50": 0,
    }

    for inn, rows in inn_rows.items():
        for i, r in enumerate(rows):
            ws.cell(r, 19).value = 1 if i == 0 else 0

    for r in range(5, max_row + 1):
        name = ws.cell(r, 2).value
        inn = str(ws.cell(r, 4).value or "").strip()
        if name is None and not inn:
            continue
        stats["rows"] += 1
        if not inn:
            continue
        stats["with_inn"] += 1
        unique = ws.cell(r, 19).value == 1

        # Enrich SSCHR from FNS open data (authoritative for this model)
        if inn in sschr:
            headcount, date_sost = sschr[inn]
            year = date_sost[-4:] if date_sost else "2025"
            ws.cell(r, 7).value = headcount
            ws.cell(r, 8).value = int(year) if year.isdigit() else year
            ws.cell(r, 17).value = (
                f"ФНС открытые данные ССЧР (ДатаСост {date_sost}; набор 7707329152-sshr2019)"
            )
            style_input(ws.cell(r, 7))
            style_input(ws.cell(r, 8))

        # Enrich OKVED
        if inn in okved and okved[inn].get("main"):
            item = okved[inn]
            ws.cell(r, 9).value = item["main"]
            all_codes = [item["main"]] + list(item.get("all") or [])
            # unique preserve
            seen = []
            for c in all_codes:
                if c not in seen:
                    seen.append(c)
            ws.cell(r, 10).value = "; ".join(seen)
            ws.cell(r, 18).value = "ЕГРЮЛ / egrul.nalog.ru (выписка)"
            style_input(ws.cell(r, 9))
            style_input(ws.cell(r, 10))

        g = ws.cell(r, 7).value
        i_val = ws.cell(r, 9).value
        j_val = ws.cell(r, 10).value

        # Formulas / computed classification (values for speed + filter sheets)
        if g in (None, ""):
            ws.cell(r, 11).value = ""
            ws.cell(r, 23).value = ""
            ge22 = False
        else:
            try:
                ge22 = float(g) >= 22
            except Exception:
                ge22 = False
            ws.cell(r, 11).value = "Да" if ge22 else "Нет"
            ws.cell(r, 23).value = "Да" if ge22 else "Нет"

        target = is_target_okved(str(i_val) if i_val else None, str(j_val) if j_val else None)
        ws.cell(r, 12).value = target
        ws.cell(r, 22).value = target

        pkg, upsell = package_for_headcount(g)
        # Commercial offer is always pilot 10 cabinets = 3360
        ws.cell(r, 13).value = pkg if pkg else "Пилот 10 каб."
        ws.cell(r, 14).value = 3360
        if g not in (None, ""):
            try:
                gn = float(g)
                if gn <= 15:
                    cat = "1–15"
                elif gn <= 30:
                    cat = "16–30"
                elif gn <= 50:
                    cat = "31–50"
                elif gn <= 100:
                    cat = "51–100"
                elif gn <= 200:
                    cat = "101–200"
                elif gn <= 500:
                    cat = "201–500"
                else:
                    cat = "500+"
                ws.cell(r, 15).value = cat
            except Exception:
                ws.cell(r, 15).value = ""
        else:
            ws.cell(r, 15).value = ""

        prio1 = target == "Да" and ge22
        strategic50 = (
            g not in (None, "")
            and _safe_float(g) is not None
            and _safe_float(g) >= 50
            and not prio1
        )
        hot = prio1 or strategic50
        okved_only = target == "Да" and not ge22
        ge22_only = ge22 and target != "Да"

        if prio1:
            prio = "1 — Горячий"
            seg = "ОКВЭД + 22+"
        elif strategic50:
            prio = "1 — Горячий"
            seg = "50+ стратегический"
        elif okved_only:
            prio = "2 — Целевой ОКВЭД"
            seg = "Целевой ОКВЭД"
        elif ge22_only:
            prio = "3 — 22+ без ОКВЭД"
            seg = "22+ без целевого ОКВЭД"
        elif target == "Не определено":
            prio = "4 — Не обогащён"
            seg = "Нет ОКВЭД/ССЧР"
        else:
            prio = "5 — Нецелевой"
            seg = "Нецелевой"

        ws.cell(r, 20).value = prio
        ws.cell(r, 21).value = seg
        ws.cell(r, 24).value = (
            "Да"
            if prio1
            else ("Нет" if target != "Не определено" and g not in (None, "") else "")
        )

        # Status
        has_s = g not in (None, "")
        has_o = i_val not in (None, "")
        if has_s and has_o:
            ws.cell(r, 16).value = "ССЧР + ОКВЭД обогащены"
        elif has_s:
            ws.cell(r, 16).value = "ССЧР обогащена; ОКВЭД не найден/не загружен"
        elif has_o:
            ws.cell(r, 16).value = "ОКВЭД есть; ССЧР нет в открытых данных ФНС"
        else:
            ws.cell(r, 16).value = "Требует обогащения"

        for c in (11, 12, 13, 14, 15, 20, 21, 22, 23, 24):
            style_formula(ws.cell(r, c))

        if unique:
            if has_s:
                stats["sschr"] += 1
            if has_o:
                stats["okved"] += 1
            if target == "Да":
                stats["target_okved"] += 1
            if ge22:
                stats["ge22"] += 1
            if _safe_float(g) is not None and _safe_float(g) >= 50:
                stats["ge50"] += 1
            if hot:
                stats["hot"] += 1
            if okved_only:
                stats["okved_only"] += 1
            if ge22_only:
                stats["ge22_only"] += 1

    # Freeze / filter
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:X{max_row}"
    set_col_widths(
        ws,
        {
            "A": 12,
            "B": 36,
            "C": 16,
            "D": 14,
            "E": 14,
            "F": 28,
            "G": 12,
            "H": 10,
            "I": 12,
            "J": 28,
            "K": 12,
            "L": 12,
            "M": 12,
            "N": 14,
            "O": 12,
            "P": 28,
            "Q": 28,
            "R": 22,
            "S": 10,
            "T": 16,
            "U": 18,
            "V": 12,
            "W": 12,
            "X": 14,
        },
    )
    return stats


def _safe_float(v):
    try:
        return float(v)
    except Exception:
        return None


def rebuild_instruction(wb):
    name = "00_Инструкция"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 0)
    ws["A1"] = "Инвестиционная модель проекта «1С:Кабинет сотрудника» — инструкция"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    ws["A3"] = "Назначение"
    ws["A3"].font = SECTION_FONT
    ws["A4"] = (
        "Модель для защиты инвестиций: рынок (TAM–SAM–SOM), сегментация клиентской базы, "
        "план/факт продаж, P&L и ДДС. Все вводимые параметры — только на листе «01_Исходные_данные»."
    )
    ws.merge_cells("A4:F4")

    ws["A6"] = "Цветовая легенда"
    ws["A6"].font = SECTION_FONT
    ws["A7"] = "Зелёные ячейки — ввод данных"
    style_input(ws["B7"])
    ws["C7"] = "Меняйте только их"
    ws["A8"] = "Серые ячейки — формулы"
    style_formula(ws["B8"])
    ws["C8"] = "Не редактировать"
    ws["A9"] = "Красные ячейки — контрольные показатели"
    style_control(ws["B9"])
    ws["C9"] = "KPI / точка безубыточности / алерты"

    ws["A11"] = "Что вводится вручную"
    ws["A11"].font = SECTION_FONT
    inputs = [
        "Стоимость лицензии / закупка 1С",
        "Количество менеджеров и оклады",
        "Размер премии по сценариям",
        "Плановые конверсии воронки (гипотезы)",
        "План продаж на менеджера (3 / 5 / 7)",
        "Плановые конверсии воронки",
        "Маркетинг и прочие расходы",
        "Факт воронки: обработано → переговоры → демо → сделка → оплата",
        "Уточнение ССЧР / ОКВЭД в базе (если есть свежие данные)",
    ]
    for i, t in enumerate(inputs, start=12):
        ws.cell(i, 1).value = "• " + t

    ws["A21"] = "Что считается автоматически"
    ws["A21"].font = SECTION_FONT
    autos = [
        "Конверсии этапов и общая конверсия",
        "Продажи, выручка, маржа, премии",
        "Выполнение плана менеджеров",
        "P&L и ДДС (помесячно и по сценариям)",
        "TAM / SAM / SOM",
        "Сегменты базы: ОКВЭД / 22+ / пересечение",
        "Горячие клиенты и потенциальная выручка",
        "Dashboard для инвестора",
    ]
    for i, t in enumerate(autos, start=22):
        ws.cell(i, 1).value = "• " + t

    ws["A31"] = "Порядок работы"
    ws["A31"].font = SECTION_FONT
    steps = [
        "1. Откройте «01_Исходные_данные» — проверьте тариф, оклады, премии, конверсии.",
        "2. На «05_База_клиентов» фильтруйте приоритеты; горячие автоматически на листе «06_Горячие_клиенты».",
        "3. Заполните факт на «08_Факт_воронка» (зелёные поля) — обновятся KPI, P&L, ДДС, Dashboard.",
        "4. Для инвестора используйте «11_Dashboard» и «03_TAM_SAM_SOM».",
        "5. Не копируйте конверсии вручную на другие листы — они тянутся формулами.",
    ]
    for i, t in enumerate(steps, start=32):
        ws.cell(i, 1).value = t
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)

    ws["A38"] = "Важно про качество данных"
    ws["A38"].font = SECTION_FONT
    ws["A39"] = (
        "ССЧР подтянуты из открытых данных ФНС (набор 7707329152-sshr2019, данные за 2025). "
        "ОКВЭД — из выписок ЕГРЮЛ. Покрытие неполное: ИП и часть ЮЛ могут отсутствовать в открытых наборах. "
        "Не экстраполируйте подтверждённый сегмент на всю базу без оговорок."
    )
    ws.merge_cells("A39:F41")
    ws["A39"].alignment = Alignment(wrap_text=True, vertical="top")

    set_col_widths(ws, {"A": 70, "B": 14, "C": 40, "D": 20, "E": 20, "F": 20})
    ws.row_dimensions[4].height = 35
    ws.row_dimensions[39].height = 60


def rebuild_params(wb, stats: dict):
    name = "01_Исходные_данные"
    # migrate from old name if needed
    if "01_Параметры" in wb.sheetnames and name not in wb.sheetnames:
        wb["01_Параметры"].title = name
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 1)

    ws["A1"] = "Исходные данные — единственная точка ввода параметров модели"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = "Зелёные — ввод. Серые — формулы. Красные — контрольные. Любое изменение пересчитывает всю модель."
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="666666")

    # A. Product
    ws["A4"] = "A. Продукт и unit-экономика"
    ws["A4"].font = SECTION_FONT
    rows_a = [
        (5, "КП: 10 кабинетов (пилот КЭДО), ₽/год", 3360, "₽/год", "Пилотное подключение + дальнейшее расширение"),
        (6, "Закупка 1С (партнёрская)", 1680, "₽/год", "Дилерская цена"),
        (7, "Маржинальный доход / клиент", "=B5-B6", "₽/год", "Цена − закупка"),
        (8, "Маржинальность до продажных расходов", "=IF(B5=0,0,B7/B5)", "%", "Маржа / цена"),
        (9, "Порог «горячий» по численности", 22, "чел.", "ОКВЭД + N+"),
        (10, "Порог стратегический по численности", 50, "чел.", "50+ = горячий даже без ОКВЭД"),
    ]
    for r, label, val, unit, note in rows_a:
        ws.cell(r, 1).value = label
        ws.cell(r, 2).value = val
        ws.cell(r, 3).value = unit
        ws.cell(r, 4).value = note
        if isinstance(val, str) and str(val).startswith("="):
            style_formula(ws.cell(r, 2))
        else:
            style_input(ws.cell(r, 2))
    ws["B8"].number_format = "0.0%"

    # B. Funnel hypotheses
    ws["A12"] = "B. Воронка и планы (гипотезы — не факт)"
    ws["A12"].font = SECTION_FONT
    headers = ["Показатель", "Малый", "Средний", "Перевыполнение", "Ед.", "Комментарий"]
    for c, h in enumerate(headers, 1):
        ws.cell(13, c).value = h
    style_header_row(ws, 13, 6)

    funnel = [
        (14, "Мощность обработки / менеджер / мес.", 150, 150, 150, "шт.", "Максимум, который менеджер может обработать"),
        (15, "Обработка → переговоры", 0.25, 0.30, 0.35, "%", "Плановая конверсия; факт — на листе воронки"),
        (16, "Переговоры → демо", 0.40, 0.45, 0.50, "%", "Плановая конверсия"),
        (17, "Демо → сделка", 0.45, 0.50, 0.55, "%", "Плановая конверсия"),
        (18, "Сделка → оплата", 0.80, 0.85, 0.90, "%", "Плановая конверсия"),
        (19, "Общая конверсия обработка → оплата", "=B15*B16*B17*B18", "=C15*C16*C17*C18", "=D15*D16*D17*D18", "%", "Автоматический расчёт"),
        (20, "План продаж / менеджер / месяц", 3, 5, 7, "шт.", "План: малый / средний / перевыполнение"),
        (21, "Требуемая обработка для плана", "=IF(B19=0,0,ROUNDUP(B20/B19,0))", "=IF(C19=0,0,ROUNDUP(C20/C19,0))", "=IF(D19=0,0,ROUNDUP(D20/D19,0))", "шт.", "Сколько клиентов нужно обработать при плановых конверсиях"),
        (22, "Прогноз оплат при требуемой обработке", "=B21*B19", "=C21*C19", "=D21*D19", "шт.", "Всегда ≥ плана за счёт ROUNDUP"),
        (23, "Контроль: прогноз ≥ план", '=IF(B22>=B20,"OK","МАЛО")', '=IF(C22>=C20,"OK","МАЛО")', '=IF(D22>=D20,"OK","МАЛО")', "—", "Красный = прогноз ниже плана"),
    ]
    for r, label, b, c, d, unit, note in funnel:
        ws.cell(r, 1).value = label
        ws.cell(r, 2).value = b
        ws.cell(r, 3).value = c
        ws.cell(r, 4).value = d
        ws.cell(r, 5).value = unit
        ws.cell(r, 6).value = note
        for col in (2, 3, 4):
            cell = ws.cell(r, col)
            if isinstance(cell.value, str) and str(cell.value).startswith("="):
                style_formula(cell)
                if r == 19:
                    cell.number_format = "0.00%"
                elif r in (21, 22):
                    cell.number_format = "0.00"
                elif r == 23:
                    style_control(cell)
            else:
                style_input(cell)
                if r in (15, 16, 17, 18):
                    cell.number_format = "0%"
    for col in (2, 3, 4):
        ws.cell(19, col).number_format = "0.00%"
        style_control(ws.cell(21, col))
        style_control(ws.cell(22, col))

    # Conditional formatting for extreme conversions
    ws.conditional_formatting.add(
        "B15:D18",
        CellIsRule(operator="greaterThan", formula=["0.95"], fill=RED),
    )
    ws.conditional_formatting.add(
        "B15:D18",
        CellIsRule(operator="lessThan", formula=["0.02"], fill=ORANGE),
    )
    ws.conditional_formatting.add(
        "B23:D23",
        FormulaRule(formula=['B23="МАЛО"'], fill=RED),
    )

    # C. Bonuses
    ws["A25"] = "C. Премии (% от маржинального дохода)"
    ws["A25"].font = SECTION_FONT
    for c, h in enumerate(["Показатель", "Малый", "Средний", "Перевыполнение", "Ед.", "Комментарий"], 1):
        ws.cell(26, c).value = h
    style_header_row(ws, 26, 6)
    ws["A27"] = "Премия от маржи"
    ws["B27"] = 0.20
    ws["C27"] = 0.30
    ws["D27"] = 0.35
    ws["E27"] = "%"
    ws["F27"] = "От маржинального дохода, не от выручки"
    for col in (2, 3, 4):
        style_input(ws.cell(27, col))
        ws.cell(27, col).number_format = "0%"

    # D. Team
    ws["A29"] = "D. Команда и расходы"
    ws["A29"].font = SECTION_FONT
    for c, h in enumerate(["Менеджер", "Оклад ₽/мес.", "Активен", "Примечание"], 1):
        ws.cell(30, c).value = h
    style_header_row(ws, 30, 4)
    managers = [("Оглоблина", 50400), ("Юнусова", 50400), ("Кургузов", 51200)]
    for i, (m, sal) in enumerate(managers):
        r = 31 + i
        ws.cell(r, 1).value = m
        ws.cell(r, 2).value = sal
        ws.cell(r, 3).value = "Да"
        style_input(ws.cell(r, 2))
        style_input(ws.cell(r, 3))
        ws.cell(r, 2).number_format = "#,##0"
    ws["A34"] = "Количество активных менеджеров"
    ws["B34"] = '=COUNTIF(C31:C33,"Да")'
    style_formula(ws["B34"])
    style_control(ws["B34"])
    ws["A35"] = "Сумма окладов активных"
    ws["B35"] = '=SUMIF(C31:C33,"Да",B31:B33)'
    style_formula(ws["B35"])
    ws["B35"].number_format = "#,##0"
    ws["A36"] = "Маркетинг и реклама, ₽/мес."
    ws["B36"] = 0
    style_input(ws["B36"])
    ws["A37"] = "Техподдержка / сопровождение, ₽/мес."
    ws["B37"] = 0
    style_input(ws["B37"])
    ws["A38"] = "Прочие расходы, ₽/мес."
    ws["B38"] = 0
    style_input(ws["B38"])

    # E. Market & base quality — linked stats
    ws["A40"] = "E. Рынок и качество клиентской базы (автоиз базы + ручные рыночные вводы)"
    ws["A40"].font = SECTION_FONT
    for c, h in enumerate(["Показатель", "Значение", "Ед.", "Источник / статус"], 1):
        ws.cell(41, c).value = h
    style_header_row(ws, 41, 4)

    market_rows = [
        (42, "Все организации РФ (ориентир)", 3_139_000, "шт.", "Верхняя граница; НЕ целевой TAM продукта", False),
        (43, "Записей с ИНН в нашей базе", stats.get("with_inn", 0), "шт.", "Факт базы", True),
        (44, "Уникальных ИНН", stats.get("unique_inn", 0), "шт.", "Дедупликация", True),
        (45, "Уникальных ИНН с ССЧР", stats.get("sschr", 0), "шт.", "ФНС открытые данные", True),
        (46, "Уникальных ИНН с ОКВЭД", stats.get("okved", 0), "шт.", "ЕГРЮЛ", True),
        (47, "Целевой ОКВЭД (подтверждено)", stats.get("target_okved", 0), "шт.", "По списку ICP", True),
        (48, "Численность ≥22", stats.get("ge22", 0), "шт.", "Порог из B9", True),
        (49, "Численность ≥50", stats.get("ge50", 0), "шт.", "Порог из B10", True),
        (50, "Горячие к обработке (ОКВЭД+22+ или 50+)", 0, "шт.", "Лист «Горячие клиенты»", True),
        (51, "Только целевой ОКВЭД (без 22+)", 0, "шт.", "Сегмент 2", True),
        (52, "Только 22+ без целевого ОКВЭД", 0, "шт.", "Сегмент 3", True),
        (53, "ИТОГО подтверждённая целевая база", 0, "шт.", "Горячие ∪ сегмент2 ∪ сегмент3", True),
        (54, "Покрытие ССЧР", None, "доля", "ССЧР / уникальные ИНН", True),
        (55, "Покрытие ОКВЭД", None, "доля", "ОКВЭД / уникальные ИНН", True),
        (56, "SAM: компании РФ с целевыми ОКВЭД", "", "шт.", "Введите после федерального среза", False),
        (57, "SAM: целевые ОКВЭД + 22+", "", "шт.", "Введите после cross-join", False),
    ]

    market_values = {
        42: 3_139_000,
        43: stats.get("with_inn", 0),
        44: stats.get("unique_inn", 0),
        45: stats.get("sschr", 0),
        46: stats.get("okved", 0),
        47: stats.get("target_okved", 0),
        48: stats.get("ge22", 0),
        49: stats.get("ge50", 0),
        50: stats.get("hot", 0),
        51: stats.get("okved_only", 0),
        52: stats.get("ge22_only", 0),
        53: stats.get("target_base", 0),
        54: "=IF(B44=0,0,B45/B44)",
        55: "=IF(B44=0,0,B46/B44)",
        56: "",
        57: "",
    }

    for r, label, _default, unit, note, is_calc in market_rows:
        ws.cell(r, 1).value = label
        val = market_values.get(r, _default)
        if r in (56, 57):
            ws.cell(r, 2).value = ""
            style_input(ws.cell(r, 2))
        elif isinstance(val, str) and val.startswith("="):
            ws.cell(r, 2).value = val
            style_formula(ws.cell(r, 2))
            if r in (54, 55):
                ws.cell(r, 2).number_format = "0.00%"
        else:
            ws.cell(r, 2).value = val
            if is_calc:
                style_formula(ws.cell(r, 2))
            else:
                style_input(ws.cell(r, 2))
            if isinstance(val, (int, float)) and r not in (54, 55):
                ws.cell(r, 2).number_format = "#,##0"
        ws.cell(r, 3).value = unit
        ws.cell(r, 4).value = note

    style_control(ws["B50"])
    style_control(ws["B53"])

    # F. Target OKVED dictionary
    ws["A59"] = "F. Справочник целевых ОКВЭД (ICP)"
    ws["A59"].font = SECTION_FONT
    ws["A60"] = "Префикс"
    ws["B60"] = "Сегмент"
    style_header_row(ws, 60, 2)
    for i, (pref, seg) in enumerate(TARGET_OKVED_PREFIXES):
        ws.cell(61 + i, 1).value = pref
        ws.cell(61 + i, 2).value = seg
        style_input(ws.cell(61 + i, 1))

    set_col_widths(ws, {"A": 48, "B": 16, "C": 14, "D": 14, "E": 10, "F": 42})
    ws.freeze_panes = "A5"


def rebuild_tariffs(wb):
    name = "02_Тарифы"
    if "Тарифы" in wb.sheetnames:
        del wb["Тарифы"]
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 2)
    ws["A1"] = "Тарифы «1С:Кабинет сотрудника»"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Источник: прайс партнёра (розница / дилер). КП модели: 10 кабинетов для пилота КЭДО = 3 360 ₽/год."
    headers = ["Пакет", "Сотрудников до", "Цена ₽/мес.", "Цена ₽/год", "Закупка ₽/год", "Маржа ₽/год", "Маржинальность", "Стартовый?"]
    for c, h in enumerate(headers, 1):
        ws.cell(4, c).value = h
    style_header_row(ws, 4, 8)
    for i, (pkg, emp, month, year, cost) in enumerate(TARIFFS):
        r = 5 + i
        ws.cell(r, 1).value = pkg
        ws.cell(r, 2).value = emp
        ws.cell(r, 3).value = month
        ws.cell(r, 4).value = year
        ws.cell(r, 5).value = cost
        ws.cell(r, 6).value = f"=D{r}-E{r}"
        ws.cell(r, 7).value = f"=IF(D{r}=0,0,F{r}/D{r})"
        ws.cell(r, 8).value = "Да" if emp == 10 else "Нет"
        for c in range(1, 6):
            style_input(ws.cell(r, c))
        style_formula(ws.cell(r, 6))
        style_formula(ws.cell(r, 7))
        ws.cell(r, 7).number_format = "0.0%"
        for c in range(3, 7):
            ws.cell(r, c).number_format = '#,##0'
    ws["A14"] = "Коммерческое предложение: 10 кабинетов для пилотного подключения КЭДО — 3 360 ₽/год (закупка 1 680 ₽). После пилота — расширение по численности."
    ws["A15"] = "В плане, прогнозе, P&L и ДДС везде сумма пилота 3 360 ₽. Пакеты выше — только оценка апселла после пилота."
    set_col_widths(ws, {"A": 14, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 12})


def rebuild_market(wb):
    name = "03_TAM_SAM_SOM"
    if "07_TAM_SAM_SOM" in wb.sheetnames:
        del wb["07_TAM_SAM_SOM"]
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 3)
    P = "01_Исходные_данные"
    ws["A1"] = "TAM — SAM — SOM: рынок для инвестора"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Общий рынок организаций РФ ≠ целевой рынок. SAM заполняется после федерального среза по ОКВЭД. SOM — capacity команды."

    headers = ["Показатель", "Организаций", "Годовой чек ₽", "Объём ₽/год", "Доля", "Источник / статус"]
    for c, h in enumerate(headers, 1):
        ws.cell(4, c).value = h
    style_header_row(ws, 4, 6)

    ws["A5"] = "TAM (верхняя граница: все орг. РФ)"
    ws["B5"] = f"='{P}'!B42"
    ws["C5"] = f"='{P}'!B5"
    ws["D5"] = "=B5*C5"
    ws["E5"] = 1
    ws["F5"] = "Ориентир; не доказанный TAM продукта КЭДО"
    ws["A6"] = "SAM — целевые ОКВЭД"
    ws["B6"] = f"='{P}'!B56"
    ws["C6"] = f"='{P}'!B5"
    ws["D6"] = '=IF(B6="","" ,B6*C6)'
    ws["E6"] = '=IF(OR(B6="",B5=0),"",B6/B5)'
    ws["F6"] = "Введите число в Исходных данных после cross-join"
    ws["A7"] = "SAM — целевые ОКВЭД + 22+"
    ws["B7"] = f"='{P}'!B57"
    ws["C7"] = f"='{P}'!B5"
    ws["D7"] = '=IF(B7="","" ,B7*C7)'
    ws["E7"] = '=IF(OR(B7="",B5=0),"",B7/B5)'
    ws["F7"] = "Узкий SAM по ICP"
    ws["A8"] = "Подтверждённая целевая база (наша)"
    ws["B8"] = f"='{P}'!B53"
    ws["C8"] = f"='{P}'!B5"
    ws["D8"] = "=B8*C8"
    ws["E8"] = '=IF(B5=0,0,B8/B5)'
    ws["F8"] = "Факт нашей базы после обогащения"
    for r in range(5, 9):
        for c in range(2, 6):
            style_formula(ws.cell(r, c))
        ws.cell(r, 2).number_format = '#,##0'
        ws.cell(r, 3).number_format = '#,##0'
        ws.cell(r, 4).number_format = '#,##0'
        ws.cell(r, 5).number_format = "0.00%"

    ws["A10"] = "SOM — capacity-based (мощность команды)"
    ws["A10"].font = SECTION_FONT
    for c, h in enumerate(
        ["Сценарий", "Менеджеров", "Продаж/мен/мес", "Продаж/год", "Выручка/год", "Доля от TAM", "Комментарий"], 1
    ):
        ws.cell(11, c).value = h
    style_header_row(ws, 11, 7)
    for i, scen in enumerate(["Малый", "Средний", "Перевыполнение"]):
        r = 12 + i
        col = "BCD"[i]
        ws.cell(r, 1).value = scen
        ws.cell(r, 2).value = f"='{P}'!B34"
        ws.cell(r, 3).value = f"='{P}'!{col}20"
        ws.cell(r, 4).value = f"=B{r}*C{r}*12"
        ws.cell(r, 5).value = f"=D{r}*'{{P}}'!B5".replace("{P}", P)
        ws.cell(r, 6).value = f"=IF(D5=0,0,E{r}/D5)"
        ws.cell(r, 7).value = "Ограничение — мощность команды"
        for c in range(2, 7):
            style_formula(ws.cell(r, c))
        ws.cell(r, 6).number_format = "0.000%"
        for c in (4, 5):
            ws.cell(r, c).number_format = '#,##0'

    ws["A16"] = "SOM — market-based (если заполнен SAM)"
    ws["A16"].font = SECTION_FONT
    for c, h in enumerate(["Доля проникновения", "SAM ₽/год", "SOM ₽/год", "SOM клиентов", "Смысл"], 1):
        ws.cell(17, c).value = h
    style_header_row(ws, 17, 5)
    for i, share in enumerate([0.005, 0.01, 0.02]):
        r = 18 + i
        ws.cell(r, 1).value = share
        style_input(ws.cell(r, 1))
        ws.cell(r, 1).number_format = "0.0%"
        ws.cell(r, 2).value = "=D6"
        ws.cell(r, 3).value = f'=IF(B{r}="","",B{r}*A{r})'
        ws.cell(r, 4).value = f'=IF(OR(C{r}="",\'{P}\'!B5=0),"",C{r}/\'{P}\'!B5)'
        ws.cell(r, 5).value = "Сценарная доля SAM"
        for c in range(2, 5):
            style_formula(ws.cell(r, c))
            ws.cell(r, c).number_format = '#,##0'

    ws["A22"] = "Как читать для инвестора"
    ws["A22"].font = SECTION_FONT
    ws["A23"] = "TAM — потолок рынка. SAM — компании с нужным ОКВЭД/ICP. SOM — что реально взять мощностью 3 менеджеров при плане 3/5/7."
    ws["A24"] = "Пока SAM пуст — опирайтесь на capacity-based SOM и подтверждённую целевую базу (строка 8)."
    set_col_widths(ws, {"A": 40, "B": 14, "C": 14, "D": 14, "E": 14, "F": 36, "G": 28})


def rebuild_base_summary(wb, stats: dict):
    name = "04_Сегменты_базы"
    if "02_База и приоритеты" in wb.sheetnames:
        del wb["02_База и приоритеты"]
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 4)
    P = "01_Исходные_данные"
    ws["A1"] = "Анализ клиентской базы — три сегмента"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "1) Целевой ОКВЭД. 2) Численность 22+. 3) Пересечение ОКВЭД + 22+ (приоритет №1). Сегменты взаимоисключающие в приоритетах."

    headers = [
        "Приоритет",
        "Критерий",
        "Уникальных ИНН",
        "Доля базы",
        "Потенциал пилота ₽/год (10 каб. = 3360)",
        "Потенциал по пакетам (если есть ССЧР)",
        "Действие",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(4, c).value = h
    style_header_row(ws, 4, 7)

    rows = [
        ("1 — Горячие", "ОКВЭД + 22+ или 50+", f"='{P}'!B50", "Обрабатывать первыми; персональный оффер"),
        ("2 — Целевой ОКВЭД", "Целевой ОКВЭД, 22+ не подтверждён", f"='{P}'!B51", "Обогащать ССЧР и запускать продажи"),
        ("3 — Дополнительный", "22+ без целевого ОКВЭД", f"='{P}'!B52", "Тестировать отдельным сценарием"),
        ("ИТОГО целевая", "Объединение сегментов", f"='{P}'!B53", "Подтверждённый минимум для плана"),
    ]
    for i, (prio, crit, formula, action) in enumerate(rows):
        r = 5 + i
        ws.cell(r, 1).value = prio
        ws.cell(r, 2).value = crit
        ws.cell(r, 3).value = formula
        ws.cell(r, 4).value = f"=IF('{P}'!B44=0,0,C{r}/'{P}'!B44)"
        ws.cell(r, 5).value = f"=C{r}*'{P}'!B5"
        ws.cell(r, 6).value = ""  # filled below with computed package potential if available
        ws.cell(r, 7).value = action
        for c in range(3, 6):
            style_formula(ws.cell(r, c))
        ws.cell(r, 4).number_format = "0.00%"
        ws.cell(r, 5).number_format = '#,##0'
        if i == 0:
            style_control(ws.cell(r, 3))

    # package potentials from stats
    ws["F5"] = stats.get("pot_hot", 0)
    ws["F6"] = stats.get("pot_okved_only", 0)
    ws["F7"] = stats.get("pot_ge22_only", 0)
    ws["F8"] = "=SUM(F5:F7)"
    for r in range(5, 9):
        style_formula(ws.cell(r, 6))
        ws.cell(r, 6).number_format = '#,##0'

    ws["A10"] = "Контроль качества обогащения"
    ws["A10"].font = SECTION_FONT
    ws["A11"] = "Покрытие ССЧР"
    ws["B11"] = f"='{P}'!B54"
    ws["A12"] = "Покрытие ОКВЭД"
    ws["B12"] = f"='{P}'!B55"
    style_control(ws["B11"])
    style_control(ws["B12"])
    ws["B11"].number_format = "0.00%"
    ws["B12"].number_format = "0.00%"
    ws["A13"] = (
        "Если покрытие низкое — в инвестпрезентации показывайте подтверждённый минимум и отдельно план обогащения, "
        "а не экстраполяцию на все уникальные ИНН."
    )
    ws.merge_cells("A13:G14")
    ws["A13"].alignment = Alignment(wrap_text=True)

    set_col_widths(ws, {"A": 20, "B": 36, "C": 14, "D": 12, "E": 18, "F": 22, "G": 40})


def rebuild_hot_clients(wb):
    name = "06_Горячие_клиенты"
    if name in wb.sheetnames:
        del wb[name]
    # place after base sheet index later
    ws = wb.create_sheet(name)
    ws["A1"] = "Горячие клиенты — ежедневная работа отдела продаж"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        "Критерии: целевой ОКВЭД + 22+ сотрудников, либо 50+, либо приоритет «1 — Горячий». "
        "Заполните ответственного и статусы (зелёные)."
    )

    headers = [
        "ИНН",
        "Компания",
        "Основной ОКВЭД",
        "ССЧР",
        "Приоритет",
        "Потенциал ₽/год",
        "Ответственный менеджер",
        "Дата обработки",
        "Статус",
        "Следующий контакт",
        "Комментарий",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(4, c).value = h
    style_header_row(ws, 4, 11)

    src = wb["Данные клиентов"]
    out_r = 5
    managers = ["Оглоблина", "Юнусова", "Кургузов"]
    mi = 0
    for r in range(5, src.max_row + 1):
        if src.cell(r, 19).value != 1:
            continue
        prio = str(src.cell(r, 20).value or "")
        if not prio.startswith("1"):
            continue
        ws.cell(out_r, 1).value = src.cell(r, 4).value
        ws.cell(out_r, 2).value = src.cell(r, 2).value
        ws.cell(out_r, 3).value = src.cell(r, 9).value
        ws.cell(out_r, 4).value = src.cell(r, 7).value
        ws.cell(out_r, 5).value = src.cell(r, 20).value
        ws.cell(out_r, 6).value = src.cell(r, 14).value
        ws.cell(out_r, 7).value = managers[mi % 3]
        style_input(ws.cell(out_r, 7))
        style_input(ws.cell(out_r, 8))
        ws.cell(out_r, 9).value = "В работе"
        style_input(ws.cell(out_r, 9))
        style_input(ws.cell(out_r, 10))
        ws.cell(out_r, 11).value = src.cell(r, 6).value
        style_input(ws.cell(out_r, 11))
        for c in range(1, 7):
            style_formula(ws.cell(out_r, c))
        if ws.cell(out_r, 6).value not in (None, ""):
            ws.cell(out_r, 6).number_format = '#,##0'
        out_r += 1
        mi += 1

    ws.cell(2, 5).value = f"Записей: {out_r - 5}"
    style_control(ws.cell(2, 5))
    ws.freeze_panes = "A5"
    if out_r > 5:
        ws.auto_filter.ref = f"A4:K{out_r - 1}"
    dv = DataValidation(type="list", formula1='"Оглоблина,Юнусова,Кургузов"', allow_blank=True)
    ws.add_data_validation(dv)
    if out_r > 5:
        dv.add(f"G5:G{out_r - 1}")
    dv2 = DataValidation(
        type="list",
        formula1='"Новый,В работе,Переговоры,Демо,Счёт,Оплата,Отказ,Отложено"',
        allow_blank=True,
    )
    ws.add_data_validation(dv2)
    if out_r > 5:
        dv2.add(f"I5:I{out_r - 1}")
    set_col_widths(
        ws,
        {
            "A": 14,
            "B": 36,
            "C": 12,
            "D": 10,
            "E": 16,
            "F": 14,
            "G": 18,
            "H": 14,
            "I": 14,
            "J": 14,
            "K": 28,
        },
    )
    return out_r - 5



def rebuild_sales_plan(wb):
    name = "07_План_продаж"
    for old in ("03_План продаж", name):
        if old in wb.sheetnames:
            del wb[old]
    ws = wb.create_sheet(name)
    P = "01_Исходные_данные"
    ws["A1"] = "План продаж 3 / 5 / 7 — от базы к обработке, прогнозу и плану"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        "КП: 10 кабинетов = 3 360 ₽/год. Прогноз строится от плановых конверсий и требуемого объёма обработки; "
        "прогноз оплат всегда ≥ плана."
    )

    # ---- Block 1: base available ----
    ws["A4"] = "1. Сколько клиентов есть в базе"
    ws["A4"].font = SECTION_FONT
    for c, h in enumerate(["Сегмент", "Уникальных ИНН", "Потенциал пилота ₽/год", "Комментарий"], 1):
        ws.cell(5, c).value = h
    style_header_row(ws, 5, 4)
    base_rows = [
        (6, "Горячие (ОКВЭД+22+ или 50+)", f"='{P}'!B50", f"='{P}'!B50*'{P}'!B5", "Брать в работу первыми"),
        (7, "Целевой ОКВЭД без 22+", f"='{P}'!B51", f"='{P}'!B51*'{P}'!B5", "Обогащать ССЧР"),
        (8, "22+ без целевого ОКВЭД", f"='{P}'!B52", f"='{P}'!B52*'{P}'!B5", "Дополнительный сегмент"),
        (9, "ИТОГО целевая база", f"='{P}'!B53", f"='{P}'!B53*'{P}'!B5", "Подтверждённый минимум"),
        (10, "Вся база (уникальные ИНН)", f"='{P}'!B44", f"='{P}'!B44*'{P}'!B5", "Включая необогащённых"),
    ]
    for r, label, cnt, pot, note in base_rows:
        ws.cell(r, 1).value = label
        ws.cell(r, 2).value = cnt
        ws.cell(r, 3).value = pot
        ws.cell(r, 4).value = note
        style_formula(ws.cell(r, 2))
        style_formula(ws.cell(r, 3))
        ws.cell(r, 2).number_format = "#,##0"
        ws.cell(r, 3).number_format = "#,##0"
    style_control(ws["B6"])
    style_control(ws["B9"])

    # ---- Block 2: reverse funnel / required processing ----
    ws["A12"] = "2. Сколько нужно обработать при плановых конверсиях, чтобы выполнить план 3 / 5 / 7"
    ws["A12"].font = SECTION_FONT
    headers = [
        "Сценарий",
        "План оплат / мен / мес",
        "Общая конв. (плановая)",
        "Требуемая обработка / мен",
        "Обработка команды / мес",
        "Прогноз оплат / мен",
        "Прогноз оплат команды",
        "Прогноз ≥ план?",
        "Выручка прогноза / мен (команда)",
        "Маржа прогноза / мен (команда)",
        "Запас базы, мес. (горячие)",
        "Запас базы, мес. (целевая)",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(13, c).value = h
    style_header_row(ws, 13, 12)

    for i, scen in enumerate(["Малый", "Средний", "Перевыполнение"]):
        r = 14 + i
        col = "BCD"[i]
        ws.cell(r, 1).value = scen
        ws.cell(r, 2).value = f"='{P}'!{col}20"  # plan
        ws.cell(r, 3).value = f"='{P}'!{col}19"  # overall conv
        ws.cell(r, 4).value = f"='{P}'!{col}21"  # required processed
        ws.cell(r, 5).value = f"=D{r}*'{{P}}'!B34".replace("{P}", P)
        ws.cell(r, 6).value = f"='{P}'!{col}22"  # forecast per mgr
        ws.cell(r, 7).value = f"=F{r}*'{{P}}'!B34".replace("{P}", P)
        ws.cell(r, 8).value = f"='{P}'!{col}23"
        ws.cell(r, 9).value = f"=G{r}*'{{P}}'!B5".replace("{P}", P)
        ws.cell(r, 10).value = f"=G{r}*'{{P}}'!B7".replace("{P}", P)
        ws.cell(r, 11).value = f"=IF(E{r}=0,\"\",ROUND('{P}'!B50/E{r},1))"
        ws.cell(r, 12).value = f"=IF(E{r}=0,\"\",ROUND('{P}'!B53/E{r},1))"
        for c in range(2, 13):
            style_formula(ws.cell(r, c))
        ws.cell(r, 3).number_format = "0.00%"
        for c in (2, 4, 5, 6, 7, 11, 12):
            ws.cell(r, c).number_format = "0.0"
        for c in (9, 10):
            ws.cell(r, c).number_format = "#,##0"
        style_control(ws.cell(r, 8))
        if i == 1:
            style_control(ws.cell(r, 4))
            style_control(ws.cell(r, 7))

    ws["A18"] = (
        "Как читать: при средних плановых конверсиях, чтобы сделать план 5 оплат, менеджер должен обработать "
        "ROUNDUP(5 / общая_конверсия) клиентов. Прогноз = обработка × конверсия ≥ 5."
    )
    ws.merge_cells("A18:L18")

    # ---- Block 3: reverse stage cascade ----
    ws["A20"] = "3. Обратная воронка от плана (сколько нужно на каждом этапе)"
    ws["A20"].font = SECTION_FONT
    for c, h in enumerate(
        ["Сценарий", "Оплаты (план)", "← Сделки", "← Демо", "← Переговоры", "← Обработано", "Контроль оплат из обработки"],
        1,
    ):
        ws.cell(21, c).value = h
    style_header_row(ws, 21, 7)
    for i in range(3):
        r = 22 + i
        col = "BCD"[i]
        src = 14 + i
        ws.cell(r, 1).value = f"=A{src}"
        ws.cell(r, 2).value = f"=B{src}"  # plan payments
        # deals = payments / (deal->pay)
        ws.cell(r, 3).value = f"=IF('{P}'!{col}18=0,0,ROUNDUP(B{r}/'{P}'!{col}18,0))"
        ws.cell(r, 4).value = f"=IF('{P}'!{col}17=0,0,ROUNDUP(C{r}/'{P}'!{col}17,0))"
        ws.cell(r, 5).value = f"=IF('{P}'!{col}16=0,0,ROUNDUP(D{r}/'{P}'!{col}16,0))"
        ws.cell(r, 6).value = f"=IF('{P}'!{col}15=0,0,ROUNDUP(E{r}/'{P}'!{col}15,0))"
        ws.cell(r, 7).value = f"=F{r}*'{{P}}'!{col}19".replace("{P}", P)
        for c in range(1, 8):
            style_formula(ws.cell(r, c))
        for c in range(2, 8):
            ws.cell(r, c).number_format = "0.00"
        style_control(ws.cell(r, 7))

    # ---- Block 4: comparison forecast vs plan + finance ----
    ws["A26"] = "4. Сравнение прогноза и плана + экономика (КП 3 360 ₽)"
    ws["A26"].font = SECTION_FONT
    for c, h in enumerate(
        [
            "Сценарий",
            "Менеджеров",
            "План оплат команды",
            "Прогноз оплат команды",
            "Дельта прогноз−план",
            "Выручка по прогнозу",
            "Маржа по прогнозу",
            "Премии",
            "Оклады",
            "Результат до прочих",
        ],
        1,
    ):
        ws.cell(27, c).value = h
    style_header_row(ws, 27, 10)
    for i in range(3):
        r = 28 + i
        src = 14 + i
        col = "BCD"[i]
        ws.cell(r, 1).value = f"=A{src}"
        ws.cell(r, 2).value = f"='{P}'!B34"
        ws.cell(r, 3).value = f"=B{src}*B{r}"  # plan team
        ws.cell(r, 4).value = f"=G{src}"  # forecast team already
        ws.cell(r, 5).value = f"=D{r}-C{r}"
        ws.cell(r, 6).value = f"=D{r}*'{{P}}'!B5".replace("{P}", P)
        ws.cell(r, 7).value = f"=D{r}*'{{P}}'!B7".replace("{P}", P)
        ws.cell(r, 8).value = f"=G{r}*'{{P}}'!{col}27".replace("{P}", P)
        ws.cell(r, 9).value = f"='{P}'!B35"
        ws.cell(r, 10).value = f"=G{r}-H{r}-I{r}"
        for c in range(1, 11):
            style_formula(ws.cell(r, c))
        for c in range(3, 11):
            ws.cell(r, c).number_format = "#,##0"
        style_control(ws.cell(r, 5))
        if i == 1:
            style_control(ws.cell(r, 10))

    # Keep aliases used by PnL/DDS/Dashboard: rows 11-13 historically were team plan.
    # Map C11:H13 style used elsewhere -> use rows 28-30 going forward.
    # Update other sheets separately to point to new cells.

    ws["A32"] = "5. Воронка по источникам (модель загрузки базы)"
    ws["A32"].font = SECTION_FONT
    for c, h in enumerate(
        ["Источник", "Обработано", "Переговоры", "Демо", "Сделки", "Оплаты", "Общая конв.", "Выручка пилота", "Маржа"],
        1,
    ):
        ws.cell(33, c).value = h
    style_header_row(ws, 33, 9)
    sources = ["Наша база", "ОКВЭД", "Лекторий", "Входящие", "Партнёры", "Реклама"]
    for i, src in enumerate(sources):
        r = 34 + i
        ws.cell(r, 1).value = src
        if i == 0:
            # default = required team processing for medium scenario
            ws.cell(r, 2).value = "=E15"
            style_formula(ws.cell(r, 2))
            ws.cell(r, 3).value = f"=ROUND(B{r}*'{{P}}'!C15,0)".replace("{P}", P)
            ws.cell(r, 4).value = f"=ROUND(C{r}*'{{P}}'!C16,0)".replace("{P}", P)
            ws.cell(r, 5).value = f"=ROUND(D{r}*'{{P}}'!C17,0)".replace("{P}", P)
            ws.cell(r, 6).value = f"=ROUND(E{r}*'{{P}}'!C18,0)".replace("{P}", P)
            for c in range(3, 7):
                style_formula(ws.cell(r, c))
        else:
            for c in range(2, 7):
                ws.cell(r, c).value = 0
                style_input(ws.cell(r, c))
        ws.cell(r, 7).value = f"=IF(B{r}=0,0,F{r}/B{r})"
        ws.cell(r, 8).value = f"=F{r}*'{{P}}'!B5".replace("{P}", P)
        ws.cell(r, 9).value = f"=F{r}*'{{P}}'!B7".replace("{P}", P)
        for c in range(7, 10):
            style_formula(ws.cell(r, c))
        ws.cell(r, 7).number_format = "0.0%"
        ws.cell(r, 8).number_format = "#,##0"
        ws.cell(r, 9).number_format = "#,##0"

    ws["A41"] = "Итого источники"
    ws["B41"] = "=SUM(B34:B39)"
    ws["F41"] = "=SUM(F34:F39)"
    ws["H41"] = "=SUM(H34:H39)"
    ws["I41"] = "=SUM(I34:I39)"
    for coord in ("B41", "F41", "H41", "I41"):
        style_formula(ws[coord])
    style_control(ws["F41"])

    ws["A43"] = "Стоимость продажи (средний сценарий, от прогноза команды)"
    ws["B43"] = f"=IF(D29=0,0,('{P}'!B35+'{P}'!B36+'{P}'!B37+'{P}'!B38)/D29)"
    style_control(ws["B43"])
    ws["B43"].number_format = "#,##0"
    ws["C43"] = "₽ на 1 оплату прогноза"

    # Compatibility named block for other sheets expecting old C11:H13 / I5 etc.
    # Write hidden-friendly alias rows at AA
    ws["AA1"] = "alias"
    for i in range(3):
        r = 2 + i
        src = 28 + i
        ws.cell(r, 27).value = f"=A{src}"  # AA
        ws.cell(r, 28).value = f"=B{src}"  # AB managers
        ws.cell(r, 29).value = f"=D{src}"  # AC forecast sales team
        ws.cell(r, 30).value = f"=F{src}"  # AD revenue
        ws.cell(r, 31).value = f"=G{src}"  # AE margin
        ws.cell(r, 32).value = f"=H{src}"  # AF premium
        ws.cell(r, 33).value = f"=I{src}"  # AG salaries
        ws.cell(r, 34).value = f"=J{src}"  # AH result

    set_col_widths(ws, {get_column_letter(c): 12 for c in range(1, 13)})
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["H"].width = 14


def rebuild_fact(wb):
    name = "08_Факт_воронка"
    for old in ("04_Факт и воронка", name):
        if old in wb.sheetnames:
            del wb[old]
    ws = wb.create_sheet(name)
    P = "01_Исходные_данные"
    ws["A1"] = "Факт продаж и воронка"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Вводите только зелёные поля. Конверсии, выполнение плана, маржа — считаются автоматически и уходят в P&L / ДДС / Dashboard."

    headers = [
        "Менеджер",
        "Сценарий плана",
        "План продаж",
        "Обработано",
        "Переговоры",
        "Демо",
        "Сделки",
        "Оплаты",
        "Конв. → перег.",
        "Конв. → демо",
        "Конв. → сделка",
        "Конв. → оплата",
        "Общая конв.",
        "Выполнение плана",
        "Выручка",
        "Маржа",
        "Премия",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(4, c).value = h
    style_header_row(ws, 4, 17)

    for i, mgr in enumerate(["Оглоблина", "Юнусова", "Кургузов"]):
        r = 5 + i
        ws.cell(r, 1).value = f"='{P}'!A{31+i}"
        ws.cell(r, 2).value = "Средний"
        style_input(ws.cell(r, 2))
        ws.cell(r, 3).value = (
            f'=IF(B{r}="Малый",\'{P}\'!$B$20,IF(B{r}="Средний",\'{P}\'!$C$20,\'{P}\'!$D$20))'
        )
        style_formula(ws.cell(r, 3))
        for c in range(4, 9):
            ws.cell(r, c).value = 0
            style_input(ws.cell(r, c))
        ws.cell(r, 9).value = f"=IF(D{r}=0,0,E{r}/D{r})"
        ws.cell(r, 10).value = f"=IF(E{r}=0,0,F{r}/E{r})"
        ws.cell(r, 11).value = f"=IF(F{r}=0,0,G{r}/F{r})"
        ws.cell(r, 12).value = f"=IF(G{r}=0,0,H{r}/G{r})"
        ws.cell(r, 13).value = f"=IF(D{r}=0,0,H{r}/D{r})"
        ws.cell(r, 14).value = f"=IF(C{r}=0,0,H{r}/C{r})"
        ws.cell(r, 15).value = f"=H{r}*'{{P}}'!$B$5".replace("{P}", P)
        ws.cell(r, 16).value = f"=H{r}*'{{P}}'!$B$7".replace("{P}", P)
        ws.cell(r, 17).value = (
            f'=P{r}*IF(B{r}="Малый",\'{P}\'!$B$27,IF(B{r}="Средний",\'{P}\'!$C$27,\'{P}\'!$D$27))'
        )
        for c in range(9, 18):
            style_formula(ws.cell(r, c))
        for c in range(9, 15):
            ws.cell(r, c).number_format = "0.0%"
        for c in (15, 16, 17):
            ws.cell(r, c).number_format = '#,##0'

    dv = DataValidation(type="list", formula1='"Малый,Средний,Перевыполнение"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add("B5:B7")

    # totals
    ws["A8"] = "ИТОГО"
    ws["C8"] = "=SUM(C5:C7)"
    for c, col in enumerate(["D", "E", "F", "G", "H"], 4):
        ws.cell(8, c).value = f"=SUM({col}5:{col}7)"
    ws["I8"] = "=IF(D8=0,0,E8/D8)"
    ws["J8"] = "=IF(E8=0,0,F8/E8)"
    ws["K8"] = "=IF(F8=0,0,G8/F8)"
    ws["L8"] = "=IF(G8=0,0,H8/G8)"
    ws["M8"] = "=IF(D8=0,0,H8/D8)"
    ws["N8"] = "=IF(C8=0,0,H8/C8)"
    ws["O8"] = "=SUM(O5:O7)"
    ws["P8"] = "=SUM(P5:P7)"
    ws["Q8"] = "=SUM(Q5:Q7)"
    for c in range(3, 18):
        style_formula(ws.cell(8, c))
    style_control(ws["N8"])
    style_control(ws["H8"])

    # sources fact
    ws["A10"] = "Факт по источникам"
    ws["A10"].font = SECTION_FONT
    for c, h in enumerate(
        ["Источник", "Обработано", "Переговоры", "Демо", "Сделки", "Оплаты", "Общая конв.", "Выручка", "Маржа"],
        1,
    ):
        ws.cell(11, c).value = h
    style_header_row(ws, 11, 9)
    for i, src in enumerate(["Наша база", "ОКВЭД", "Лекторий", "Входящие", "Партнёры", "Реклама"]):
        r = 12 + i
        ws.cell(r, 1).value = src
        for c in range(2, 7):
            ws.cell(r, c).value = 0
            style_input(ws.cell(r, c))
        ws.cell(r, 7).value = f"=IF(B{r}=0,0,F{r}/B{r})"
        ws.cell(r, 8).value = f"=F{r}*'{{P}}'!$B$5".replace("{P}", P)
        ws.cell(r, 9).value = f"=F{r}*'{{P}}'!$B$7".replace("{P}", P)
        for c in range(7, 10):
            style_formula(ws.cell(r, c))
        ws.cell(r, 7).number_format = "0.0%"

    ws["A19"] = "Контроль: сумма оплат по менеджерам должна сходиться с суммой оплат по источникам (если ведёте оба разреза)."
    ws["A20"] = "Оплаты менеджеры"
    ws["B20"] = "=H8"
    ws["A21"] = "Оплаты источники"
    ws["B21"] = "=SUM(F12:F17)"
    ws["A22"] = "Расхождение"
    ws["B22"] = "=B20-B21"
    style_control(ws["B22"])

    set_col_widths(ws, {get_column_letter(c): 11 for c in range(1, 18)})
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12


def rebuild_pnl(wb):
    name = "09_PnL"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    P = "01_Исходные_данные"
    SP = "07_План_продаж"
    F = "08_Факт_воронка"
    ws["A1"] = "Финансовый план (P&L) — экономика проекта"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "План по сценариям и факт месяца. Налоги/взносы в текущей версии не моделируются."

    headers = ["Статья", "Малый план", "Средний план", "Перевыполнение", "Факт месяца", "Комментарий"]
    for c, h in enumerate(headers, 1):
        ws.cell(4, c).value = h
    style_header_row(ws, 4, 6)

    lines = [
        (5, "Выручка", f"='{SP}'!F28", f"='{SP}'!F29", f"='{SP}'!F30", f"='{F}'!O8", "Оплаты × цена"),
        (6, "Себестоимость (закупка 1С)", f"=-B5*'{{P}}'!B6/'{{P}}'!B5".replace("{P}", P), f"=-C5*'{{P}}'!B6/'{{P}}'!B5".replace("{P}", P), f"=-D5*'{{P}}'!B6/'{{P}}'!B5".replace("{P}", P), f"=-E5*'{{P}}'!B6/'{{P}}'!B5".replace("{P}", P), "Пропорция закупки"),
        (7, "Маржинальный доход", "=B5+B6", "=C5+C6", "=D5+D6", "=E5+E6", "После закупки"),
        (8, "ФОТ (оклады)", f"=-'{P}'!B35", f"=-'{P}'!B35", f"=-'{P}'!B35", f"=-'{P}'!B35", "Активные менеджеры"),
        (9, "Премии", f"=-'{SP}'!H28", f"=-'{SP}'!H29", f"=-'{SP}'!H30", f"=-'{F}'!Q8", "% от маржи"),
        (10, "Маркетинг", f"=-'{P}'!B36", f"=-'{P}'!B36", f"=-'{P}'!B36", f"=-'{P}'!B36", ""),
        (11, "Техподдержка", f"=-'{P}'!B37", f"=-'{P}'!B37", f"=-'{P}'!B37", f"=-'{P}'!B37", ""),
        (12, "Прочие расходы", f"=-'{P}'!B38", f"=-'{P}'!B38", f"=-'{P}'!B38", f"=-'{P}'!B38", ""),
        (13, "Операционная прибыль", "=B7+B8+B9+B10+B11+B12", "=C7+C8+C9+C10+C11+C12", "=D7+D8+D9+D10+D11+D12", "=E7+E8+E9+E10+E11+E12", "До налогов"),
        (14, "Маржинальность (маржа/выручка)", "=IF(B5=0,0,B7/B5)", "=IF(C5=0,0,C7/C5)", "=IF(D5=0,0,D7/D5)", "=IF(E5=0,0,E7/E5)", ""),
        (15, "Рентабельность (прибыль/выручка)", "=IF(B5=0,0,B13/B5)", "=IF(C5=0,0,C13/C5)", "=IF(D5=0,0,D13/D5)", "=IF(E5=0,0,E13/E5)", ""),
    ]
    for r, label, b, c, d, e, note in lines:
        ws.cell(r, 1).value = label
        ws.cell(r, 2).value = b
        ws.cell(r, 3).value = c
        ws.cell(r, 4).value = d
        ws.cell(r, 5).value = e
        ws.cell(r, 6).value = note
        for col in range(2, 6):
            style_formula(ws.cell(r, col))
            if r in (14, 15):
                ws.cell(r, col).number_format = "0.0%"
            else:
                ws.cell(r, col).number_format = '#,##0'
        if r == 13:
            for col in range(2, 6):
                style_control(ws.cell(r, col))

    set_col_widths(ws, {"A": 34, "B": 14, "C": 14, "D": 14, "E": 14, "F": 28})


def rebuild_cashflow(wb):
    name = "10_ДДС"
    for old in ("05_ДДС_АПУ", name):
        if old in wb.sheetnames:
            del wb[old]
    ws = wb.create_sheet(name)
    P = "01_Исходные_данные"
    SP = "07_План_продаж"
    F = "08_Факт_воронка"
    ws["A1"] = "ДДС проекта — когда выходим в плюс"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        "Годовая подписка поступает разово при оплате. Ниже: сценарии плана и помесячный ДДС "
        "среднего сценария на 12 месяцев."
    )

    ws["A4"] = "Сводка по сценариям (месяц в стационаре)"
    ws["A4"].font = SECTION_FONT
    for c, h in enumerate(
        [
            "Сценарий",
            "Оплат / мес.",
            "Поступления",
            "Закупка 1С",
            "Оклады",
            "Премии",
            "Прочие",
            "Денежный поток",
            "Годовой поток",
        ],
        1,
    ):
        ws.cell(5, c).value = h
    style_header_row(ws, 5, 9)
    for i, scen in enumerate(["Малый", "Средний", "Перевыполнение"]):
        r = 6 + i
        ws.cell(r, 1).value = scen
        ws.cell(r, 2).value = f"='{SP}'!D{28+i}"
        ws.cell(r, 3).value = f"=B{r}*'{{P}}'!B5".replace("{P}", P)
        ws.cell(r, 4).value = f"=B{r}*'{{P}}'!B6".replace("{P}", P)
        ws.cell(r, 5).value = f"='{P}'!B35"
        ws.cell(r, 6).value = f"='{SP}'!H{28+i}"
        ws.cell(r, 7).value = f"='{P}'!B36+'{{P}}'!B37+'{{P}}'!B38".replace("{P}", P)
        ws.cell(r, 8).value = f"=C{r}-D{r}-E{r}-F{r}-G{r}"
        ws.cell(r, 9).value = f"=H{r}*12"
        for c in range(2, 10):
            style_formula(ws.cell(r, c))
            ws.cell(r, c).number_format = "#,##0"
        if i == 1:
            style_control(ws.cell(r, 8))

    ws["A10"] = "Точка безубыточности (оплат/мес.)"
    ws["B10"] = (
        f"=IF(('{P}'!B7*(1-'{P}'!C27))=0,0,"
        f"ROUNDUP(('{P}'!B35+'{P}'!B36+'{P}'!B37+'{P}'!B38)/('{P}'!B7*(1-'{P}'!C27)),0))"
    )
    style_control(ws["B10"])
    ws["C10"] = "При средней премии 30%; без налогов"

    ws["A12"] = "Помесячный ДДС — средний сценарий (12 мес.)"
    ws["A12"].font = SECTION_FONT
    ws["A13"] = "Месяц"
    for m in range(1, 13):
        ws.cell(13, 1 + m).value = m
    style_header_row(ws, 13, 13)

    # Ramp: 40%, 60%, 80%, then 100% of steady medium plan
    ramp = [0.4, 0.6, 0.8] + [1.0] * 9
    ws["A14"] = "Коэффициент выхода"
    for m, k in enumerate(ramp, 1):
        ws.cell(14, 1 + m).value = k
        style_input(ws.cell(14, 1 + m))
        ws.cell(14, 1 + m).number_format = "0%"

    ws["A15"] = "Поступления"
    ws["A16"] = "Закупка 1С"
    ws["A17"] = "Оклады"
    ws["A18"] = "Премии"
    ws["A19"] = "Прочие расходы"
    ws["A20"] = "Денежный поток"
    ws["A21"] = "Накопленный поток"
    for m in range(1, 13):
        c = 1 + m
        col = get_column_letter(c)
        ws.cell(15, c).value = f"='{SP}'!F29*{col}14"
        ws.cell(16, c).value = f"='{SP}'!F29*'{{P}}'!B6/'{{P}}'!B5*{col}14".replace("{P}", P)
        ws.cell(17, c).value = f"='{P}'!B35"
        ws.cell(18, c).value = f"='{SP}'!H29*{col}14"
        ws.cell(19, c).value = f"='{P}'!B36+'{{P}}'!B37+'{{P}}'!B38".replace("{P}", P)
        ws.cell(20, c).value = f"={col}15-{col}16-{col}17-{col}18-{col}19"
        if m == 1:
            ws.cell(21, c).value = f"={col}20"
        else:
            prev = get_column_letter(c - 1)
            ws.cell(21, c).value = f"={prev}21+{col}20"
        for r in range(15, 22):
            style_formula(ws.cell(r, c))
            ws.cell(r, c).number_format = "#,##0"
        style_control(ws.cell(20, c))

    ws["A23"] = "Месяц выхода в плюс (накопленный поток > 0)"
    ws["B23"] = (
        '=IF(M21>0,MATCH(TRUE,INDEX(B21:M21>0,0),0),"не в горизонте 12 мес.")'
    )
    style_control(ws["B23"])

    ws["A25"] = "Факт месяца (из воронки)"
    ws["A25"].font = SECTION_FONT
    ws["A26"] = "Поступления факт"
    ws["B26"] = f"='{F}'!O8"
    ws["A27"] = "Закупка факт"
    ws["B27"] = f"='{F}'!H8*'{P}'!B6"
    ws["A28"] = "Оклады"
    ws["B28"] = f"='{P}'!B35"
    ws["A29"] = "Премии факт"
    ws["B29"] = f"='{F}'!Q8"
    ws["A30"] = "Прочие"
    ws["B30"] = f"='{P}'!B36+'{P}'!B37+'{P}'!B38"
    ws["A31"] = "Поток факт"
    ws["B31"] = "=B26-B27-B28-B29-B30"
    for r in range(26, 32):
        style_formula(ws.cell(r, 2))
        ws.cell(r, 2).number_format = "#,##0"
    style_control(ws["B31"])

    set_col_widths(ws, {"A": 28, **{get_column_letter(c): 11 for c in range(2, 14)}})


def rebuild_dashboard(wb, hot_count: int, stats: dict):
    name = "11_Dashboard"
    for old in ("00_Инвесторский дашборд", name):
        if old in wb.sheetnames:
            del wb[old]
    ws = wb.create_sheet(name, 0)
    # move instruction first later via reorder
    P = "01_Исходные_данные"
    SP = "07_План_продаж"
    F = "08_Факт_воронка"
    PN = "09_PnL"
    CF = "10_ДДС"
    MKT = "03_TAM_SAM_SOM"

    ws["A1"] = "Dashboard инвестора — «1С:Кабинет сотрудника»"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:H1")
    ws["A2"] = "Все KPI связаны с исходными данными, базой, планом и фактом. Меняйте только зелёные поля на листе исходных данных и факта."
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="666666")

    ws["A4"] = "Клиентская база"
    ws["A4"].font = SECTION_FONT
    kpis_base = [
        (5, "Уникальных клиентов", f"='{P}'!B44", "шт."),
        (6, "С целевым ОКВЭД", f"='{P}'!B47", "шт."),
        (7, "С численностью 22+", f"='{P}'!B48", "шт."),
        (8, "Горячие (приоритет №1 / 50+)", f"='{P}'!B50", "шт."),
        (9, "Горячие на листе менеджеров", hot_count, "шт."),
        (10, "Покрытие ССЧР", f"='{P}'!B54", "%"),
        (11, "Покрытие ОКВЭД", f"='{P}'!B55", "%"),
    ]
    ws["A4"] = "Показатель"
    ws["B4"] = "Значение"
    ws["C4"] = "Ед."
    style_header_row(ws, 4, 3)
    for r, label, val, unit in kpis_base:
        ws.cell(r, 1).value = label
        ws.cell(r, 2).value = val
        ws.cell(r, 3).value = unit
        style_formula(ws.cell(r, 2))
        if unit == "%":
            ws.cell(r, 2).number_format = "0.0%"
        else:
            ws.cell(r, 2).number_format = "#,##0"
    style_control(ws["B8"])

    ws["E4"] = "Продажи и финансы (средний план / факт)"
    ws["E4"].font = SECTION_FONT
    ws["E5"] = "Показатель"
    ws["F5"] = "План (средний)"
    ws["G5"] = "Факт"
    style_header_row(ws, 5, 3)
    # rewrite headers properly in E-G
    ws["E5"] = "Показатель"
    ws["F5"] = "План (средний)"
    ws["G5"] = "Факт"
    for c in range(5, 8):
        ws.cell(5, c).fill = BLUE_HDR
        ws.cell(5, c).font = WHITE_FONT

    fin = [
        (6, "План оплат", f"='{SP}'!C29", f"='{F}'!H8"),
        (7, "Прогноз оплат (≥ план)", f"='{SP}'!D29", ""),
        (8, "Выручка по прогнозу", f"='{SP}'!F29", f"='{F}'!O8"),
        (9, "Маржа по прогнозу", f"='{SP}'!G29", f"='{F}'!P8"),
        (10, "Прибыль (P&L)", f"='{PN}'!C13", f"='{PN}'!E13"),
        (11, "Выполнение плана", "", f"='{F}'!N8"),
        (12, "Общая конверсия факт", "", f"='{F}'!M8"),
        (13, "Денежный поток", f"='{CF}'!H7", f"='{CF}'!B31"),
    ]
    for r, label, plan, fact in fin:
        ws.cell(r, 5).value = label
        ws.cell(r, 6).value = plan if plan != "" else "—"
        ws.cell(r, 7).value = fact if fact != "" else "—"
        if plan != "":
            style_formula(ws.cell(r, 6))
            ws.cell(r, 6).number_format = "#,##0"
        if fact != "":
            style_formula(ws.cell(r, 7))
            if r in (11, 12):
                ws.cell(r, 7).number_format = "0.0%"
            else:
                ws.cell(r, 7).number_format = "#,##0"
    style_control(ws["G11"])
    style_control(ws["F7"])
    style_control(ws["F10"])

    ws["A13"] = "Рынок"
    ws["A13"].font = SECTION_FONT
    ws["A14"] = "TAM ₽/год"
    ws["B14"] = f"='{MKT}'!D5"
    ws["A15"] = "SAM ₽/год"
    ws["B15"] = f"='{MKT}'!D6"
    ws["A16"] = "SOM capacity (средний) ₽/год"
    ws["B16"] = f"='{MKT}'!E13"
    ws["A17"] = "Подтверждённая база ₽/год"
    ws["B17"] = f"='{MKT}'!D8"
    for r in range(14, 18):
        style_formula(ws.cell(r, 2))
        ws.cell(r, 2).number_format = "#,##0"

    ws["E14"] = "Unit-экономика"
    ws["E14"].font = SECTION_FONT
    ws["E15"] = "Цена"
    ws["F15"] = f"='{P}'!B5"
    ws["E16"] = "Маржа / клиент"
    ws["F16"] = f"='{P}'!B7"
    ws["E17"] = "Маржинальность"
    ws["F17"] = f"='{P}'!B8"
    ws["E18"] = "Безубыточность, оплат/мес"
    ws["F18"] = f"='{CF}'!B10"
    ws["E19"] = "КП пилота (10 кабинетов)"
    ws["F19"] = 3360
    ws["E20"] = "Заметка"
    ws["F20"] = "Везде в плане/прогнозе/P&L/ДДС — пилот 10 кабинетов = 3 360 ₽/год. Апселл по численности смотрите в тарифах."
    for r in range(15, 20):
        style_formula(ws.cell(r, 6))
        if r == 17:
            ws.cell(r, 6).number_format = "0.0%"
        else:
            ws.cell(r, 6).number_format = "#,##0"
    style_control(ws["F18"])
    style_control(ws["F19"])
    ws["F20"].alignment = Alignment(wrap_text=True)

    ws["A20"] = "План команды по сценариям"
    ws["A20"].font = SECTION_FONT
    for c, h in enumerate(["Сценарий", "Продаж/мес", "Выручка", "Маржа", "Опер. результат"], 1):
        ws.cell(21, c).value = h
    style_header_row(ws, 21, 5)
    for i in range(3):
        r = 22 + i
        ws.cell(r, 1).value = f"='{SP}'!A{28+i}"
        ws.cell(r, 2).value = f"='{SP}'!D{28+i}"
        ws.cell(r, 3).value = f"='{SP}'!F{28+i}"
        ws.cell(r, 4).value = f"='{SP}'!G{28+i}"
        ws.cell(r, 5).value = f"='{SP}'!J{28+i}"
        for c in range(1, 6):
            style_formula(ws.cell(r, c))
        for c in range(2, 6):
            ws.cell(r, c).number_format = "#,##0"

    chart = BarChart()
    chart.type = "col"
    chart.title = "Продажи по сценариям"
    chart.y_axis.title = "Оплат / мес."
    data = Reference(ws, min_col=2, min_row=21, max_row=24)
    cats = Reference(ws, min_col=1, min_row=22, max_row=24)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 12
    chart.height = 8
    ws.add_chart(chart, "E20")

    ws["A26"] = "Контроль качества модели"
    ws["A26"].font = SECTION_FONT
    ws["A27"] = (
        f"ССЧР обогащено у {stats.get('sschr', 0)} из {stats.get('unique_inn', 0)} уникальных ИНН; "
        f"ОКВЭД — у {stats.get('okved', 0)}. "
        "Для защиты используйте подтверждённый минимум и capacity-based SOM, пока SAM не заполнен федеральным срезом."
    )
    ws.merge_cells("A27:H28")
    ws["A27"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A27"].fill = YELLOW

    set_col_widths(
        ws,
        {"A": 34, "B": 14, "C": 10, "D": 12, "E": 28, "F": 14, "G": 14, "H": 12},
    )


def rebuild_methodology(wb):
    name = "12_Методика"
    for old in ("06_Методика", name):
        if old in wb.sheetnames:
            del wb[old]
    ws = wb.create_sheet(name)
    ws["A1"] = "Методика модели — проверяемость для инвестора"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Любое число — факт, параметр или расчёт. Гипотезы не смешиваются с фактом воронки."

    headers = ["Показатель", "Что означает", "Где брать", "Тип", "Проверка"]
    for c, h in enumerate(headers, 1):
        ws.cell(4, c).value = h
    style_header_row(ws, 4, 5)
    rows = [
        ("ИНН", "Идентификатор компании", "Исходная CRM-база", "Факт", "Уникальность"),
        ("ОКВЭД", "Отраслевой профиль", "ЕГРЮЛ / egrul.nalog.ru", "Факт", "Дата выписки"),
        ("ССЧР", "Среднесписочная численность", "ФНС open data 7707329152-sshr2019", "Факт", "Год показателя"),
        ("Целевой ОКВЭД", "ICP проекта", "Справочник на листе исходных данных", "Правило", "Префиксы ОКВЭД"),
        ("Горячий клиент", "ОКВЭД+22+ или 50+", "Расчёт", "Расчёт", "Оба критерия"),
        ("Конверсия план", "Гипотеза воронки", "01_Исходные_данные", "Гипотеза", "Не подменять фактом"),
        ("Конверсия факт", "Факт этапов", "08_Факт_воронка", "Факт", "CRM / 1С"),
        ("План 3/5/7", "План продаж менеджера", "Управленческий план", "План", "Сверять с прогнозом воронки"),
        ("Цена / закупка", "Unit-экономика", "Прайс 1С", "Параметр", "Маржа = цена − закупка"),
        ("Премия", "% от маржи", "Политика мотивации", "Параметр", "Не от выручки"),
        ("TAM/SAM/SOM", "Рынок", "Федеральные реестры + capacity", "Расчёт", "SAM ⊂ TAM; SOM ≤ capacity"),
        ("P&L / ДДС", "Экономика и кэш", "Авто из плана/факта", "Расчёт", "Сходимость оплат"),
    ]
    for i, row in enumerate(rows):
        for c, v in enumerate(row, 1):
            ws.cell(5 + i, c).value = v
            ws.cell(5 + i, c).border = THIN
    ws["A18"] = "Источники обогащения этой версии"
    ws["A18"].font = SECTION_FONT
    ws["A19"] = "1) ССЧР: https://www.nalog.gov.ru/opendata/7707329152-sshr2019/ (данные за 2025, ДатаСост 31.12.2025)."
    ws["A20"] = "2) ОКВЭД: выписки ЕГРЮЛ с egrul.nalog.ru для приоритетного сегмента 22+."
    ws["A21"] = "3) Тарифы: партнёрский прайс «1С:Кабинет сотрудника» (розница/дилер)."
    ws["A22"] = "Пересборка: python3 scripts/build_investment_model.py"
    set_col_widths(ws, {"A": 22, "B": 28, "C": 40, "D": 12, "E": 22})


def rename_client_sheet(wb):
    if "Данные клиентов" in wb.sheetnames:
        wb["Данные клиентов"].title = "05_База_клиентов"


def reorder_sheets(wb):
    order = [
        "00_Инструкция",
        "11_Dashboard",
        "01_Исходные_данные",
        "02_Тарифы",
        "03_TAM_SAM_SOM",
        "04_Сегменты_базы",
        "05_База_клиентов",
        "06_Горячие_клиенты",
        "07_План_продаж",
        "08_Факт_воронка",
        "09_PnL",
        "10_ДДС",
        "12_Методика",
    ]
    # delete obsolete sheets not in order
    for s in list(wb.sheetnames):
        if s not in order:
            del wb[s]
    for idx, name in enumerate(order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=idx - wb.sheetnames.index(name))


def compute_extra_stats(wb) -> dict:
    """Second pass after enrichment for precise segment/potential stats."""
    ws = wb["Данные клиентов"] if "Данные клиентов" in wb.sheetnames else wb["05_База_клиентов"]
    stats = {
        "prio1": 0,
        "target_base": 0,
        "pot_hot": 0,
        "pot_okved_only": 0,
        "pot_ge22_only": 0,
        "hot": 0,
        "okved_only": 0,
        "ge22_only": 0,
        "target_okved": 0,
        "ge22": 0,
        "ge50": 0,
        "sschr": 0,
        "okved": 0,
        "unique_inn": 0,
        "rows_with_inn": 0,
    }
    seen = set()
    for r in range(5, ws.max_row + 1):
        inn = str(ws.cell(r, 4).value or "").strip()
        if inn:
            stats["rows_with_inn"] += 1
        if ws.cell(r, 19).value != 1:
            continue
        if not inn or inn in seen:
            continue
        seen.add(inn)
        stats["unique_inn"] += 1
        g = _safe_float(ws.cell(r, 7).value)
        target = ws.cell(r, 12).value
        prio = str(ws.cell(r, 20).value or "")
        # Pilot KP always 3360
        pot_n = 3360
        if g is not None:
            stats["sschr"] += 1
            if g >= 22:
                stats["ge22"] += 1
            if g >= 50:
                stats["ge50"] += 1
        if ws.cell(r, 9).value not in (None, ""):
            stats["okved"] += 1
        if target == "Да":
            stats["target_okved"] += 1
        seg = str(ws.cell(r, 21).value or "")
        if prio.startswith("1"):
            stats["hot"] += 1
            stats["pot_hot"] += pot_n
            if "ОКВЭД + 22+" in seg:
                stats["prio1"] += 1
        elif prio.startswith("2"):
            stats["okved_only"] += 1
            stats["pot_okved_only"] += pot_n
        elif prio.startswith("3"):
            stats["ge22_only"] += 1
            stats["pot_ge22_only"] += pot_n
    stats["with_inn"] = stats["rows_with_inn"]
    stats["target_base"] = stats["hot"] + stats["okved_only"] + stats["ge22_only"]
    return stats


def main():
    if not SRC.exists():
        raise SystemExit(f"Source not found: {SRC}")

    print("Loading workbook...")
    # copy first to avoid mutating original during failed runs
    tmp = ROOT / "_build_model_tmp.xlsx"
    shutil.copy2(SRC, tmp)
    wb = load_workbook(tmp)

    print("Loading enrichment...")
    sschr, okved = load_enrichment()
    print(f"  SSCHR keys={len(sschr)} OKVED keys={len(okved)}")

    print("Enriching client base...")
    stats = enrich_clients(wb, sschr, okved)
    extra = compute_extra_stats(wb)
    stats.update(extra)
    # with_inn from first pass is better (all rows with inn)
    print("Stats:", json.dumps(stats, ensure_ascii=False, indent=2))

    # persist okved into data/
    DATA.mkdir(exist_ok=True)
    if okved:
        (DATA / "okved_egrul.json").write_text(
            json.dumps(okved, ensure_ascii=False), encoding="utf-8"
        )
    (DATA / "model_stats.json").write_text(
        json.dumps(stats, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print("Rebuilding sheets...")
    rebuild_instruction(wb)
    rebuild_params(wb, stats)
    rebuild_tariffs(wb)
    rebuild_market(wb)
    rebuild_base_summary(wb, stats)
    hot_count = rebuild_hot_clients(wb)
    rebuild_sales_plan(wb)
    rebuild_fact(wb)
    rebuild_pnl(wb)
    rebuild_cashflow(wb)
    rebuild_dashboard(wb, hot_count, stats)
    rebuild_methodology(wb)
    rename_client_sheet(wb)
    reorder_sheets(wb)

    print(f"Saving {OUT}...")
    wb.save(OUT)
    wb.close()
    if tmp.exists():
        tmp.unlink()
    print("Done.")
    print("Sheets:", ", ".join(load_workbook(OUT, read_only=True).sheetnames))


if __name__ == "__main__":
    main()
